import flet as ft
from gym_manager.views.module_views import ModuleView
from gym_manager.utils.database import get_db_session
from gym_manager.models.routine import Rutina
from gym_manager.controllers.routine_controller import RoutineController
import os
from pathlib import Path

class RoutinesView(ModuleView):
    def __init__(self, page: ft.Page):
        super().__init__(page, "Gestión de Rutinas")
        # Inicializar referencias
        self.new_routine_name = ft.Ref[ft.TextField]()
        self.new_routine_type = ft.Ref[ft.Dropdown]()
        self.new_routine_difficulty = ft.Ref[ft.Dropdown]()
        self.new_routine_duration = ft.Ref[ft.TextField]()
        self.new_routine_description = ft.Ref[ft.TextField]()
        
        # Inicializar controlador
        self.routine_controller = RoutineController()
        
        # FilePicker para el documento de la rutina
        self.routine_file_picker = ft.FilePicker(
            on_result=self.on_routine_file_selected
        )
        self.page.overlay.append(self.routine_file_picker)
        
        self.setup_confirm_dialog()
        self.setup_view()

    def on_routine_file_selected(self, e):
        if e.files:
            self.selected_file = e.files[0]
            self.file_selected_label.value = f"Archivo seleccionado: {self.selected_file.name}"
            self.page.update()

    def setup_view(self):
        # Título amigable
        self.welcome_title = ft.Text(
            "¡Administra y consulta tus rutinas!",
            size=28,
            weight=ft.FontWeight.BOLD,
            color=ft.colors.BLACK,
            text_align=ft.TextAlign.LEFT,
        )

        # Botón Nueva Rutina
        self.new_routine_btn = ft.ElevatedButton(
            text="Nueva Rutina",
            icon=ft.icons.FITNESS_CENTER,
            style=ft.ButtonStyle(
                bgcolor=ft.colors.BLUE,
                color=ft.colors.WHITE,
                shape=ft.RoundedRectangleBorder(radius=10),
                padding=ft.padding.symmetric(horizontal=10, vertical=8),
                text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD),
            ),
            width=150,
            on_click=self.show_new_routine_modal
        )

        # Botones de exportación
        self.export_excel_btn = ft.IconButton(
            icon=ft.icons.TABLE_VIEW,
            icon_color=ft.colors.GREEN_700,
            tooltip="Exportar a Excel",
            on_click=self.export_to_excel,
            width=48,
            height=48,
        )
        self.export_pdf_btn = ft.IconButton(
            icon=ft.icons.PICTURE_AS_PDF,
            icon_color=ft.colors.RED_700,
            tooltip="Exportar a PDF",
            on_click=self.export_to_pdf,
            width=48,
            height=48,
        )

        # Filtros de búsqueda
        self.search_field = ft.TextField(
            label="Buscar por nombre de rutina",
            prefix_icon=ft.icons.SEARCH,
            border_radius=10,
            width=320,
            height=48,
            text_size=16,
            on_change=self.apply_filters
        )

        # Contador de registros
        self.records_counter = ft.Text(
            "0 rutinas",
            size=14,
            color=ft.colors.GREY_600,
            weight=ft.FontWeight.W_500,
        )

        self.difficulty_filter = ft.Dropdown(
            label="Dificultad",
            width=200,
            options=[
                ft.dropdown.Option("Todas"),
                ft.dropdown.Option("Principiante"),
                ft.dropdown.Option("Intermedio"),
                ft.dropdown.Option("Avanzado")
            ],
            border_radius=10,
            text_size=16,
            on_change=self.apply_filters
        )

        self.clear_btn = ft.OutlinedButton(
            text="Limpiar filtros",
            icon=ft.icons.CLEAR,
            style=ft.ButtonStyle(
                color=ft.colors.GREY_700,
                shape=ft.RoundedRectangleBorder(radius=8),
                padding=ft.padding.symmetric(horizontal=18, vertical=12),
                text_style=ft.TextStyle(size=16),
            ),
            on_click=self.clear_filters
        )

        # Tabla de rutinas
        self.routines_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Nombre", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Dificultad", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Descripción", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Miembros Asignados", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Acciones", size=18, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=12,
            vertical_lines=ft.border.all(1, ft.colors.GREY_300),
            horizontal_lines=ft.border.all(1, ft.colors.GREY_300),
            column_spacing=60,
            heading_row_color=ft.colors.GREY_100,
            heading_row_height=60,
            data_row_color=ft.colors.WHITE,
            data_row_min_height=56,
        )

        # Contenedor principal
        self.content = ft.Container(
            content=ft.Column(
                controls=[
                    # Header fijo con título y botones
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                self.welcome_title,
                                ft.Row(
                                    controls=[
                                        self.new_routine_btn,
                                        self.export_excel_btn,
                                        self.export_pdf_btn,
                                    ],
                                    alignment=ft.MainAxisAlignment.END,
                                ),
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        ),
                        padding=ft.padding.only(bottom=20),
                    ),
                    # Filtros fijos
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                self.search_field,
                                self.records_counter,
                                self.difficulty_filter,
                                self.clear_btn,
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        ),
                        padding=ft.padding.only(bottom=20),
                    ),
                    # Tabla con scroll
                    ft.Container(
                        content=self.routines_table,
                        expand=True,
                        height=600,  # Altura más grande para mejor visualización
                    ),
                ],
                spacing=0,
                expand=True,
                scroll=ft.ScrollMode.AUTO,  # Scroll para toda la columna
            ),
            padding=20,
            expand=True,
        )
        
        # Modal de exportación
        self.export_dialog = ft.AlertDialog(
            title=ft.Text("Confirmar Exportación", size=22, weight=ft.FontWeight.BOLD),
            content=ft.Text(
                "¿Deseas exportar las rutinas?\n"
                "El archivo se guardará en tu carpeta de Descargas.",
                size=16
            ),
            actions=[
                ft.TextButton(
                    "Cancelar",
                    on_click=self.close_export_dialog,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.WHITE,
                        color=ft.colors.BLACK87,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
                ft.ElevatedButton(
                    "Exportar",
                    on_click=self.confirm_export,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE_900,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True,
        )
        self.page.overlay.append(self.export_dialog)
        
        # Cargar datos iniciales
        self.load_data()

    def get_content(self):
        """
        Retorna el contenido de la vista
        """
        self.page.update()
        return self.content

    def setup_confirm_dialog(self):
        self.confirm_dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Confirmar eliminación"),
            content=ft.Text("¿Estás seguro de que deseas eliminar esta rutina?"),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_confirm_dialog),
                ft.TextButton("Eliminar", on_click=self.confirm_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.confirm_dialog not in self.page.overlay:
            self.page.overlay.append(self.confirm_dialog)

    def close_confirm_dialog(self, e=None):
        self.confirm_dialog.open = False
        self.page.update()

    def confirm_delete(self, e):
        if hasattr(self, 'routine_to_delete'):
            success, message = self.routine_controller.delete_routine(self.routine_to_delete.id_rutina)
            if success:
                self.show_message("Rutina eliminada exitosamente", ft.colors.GREEN)
                self.load_data()
            else:
                self.show_message(f"Error al eliminar la rutina: {message}", ft.colors.RED)
        self.close_confirm_dialog()

    def load_data(self):
        """
        Carga los datos iniciales de la vista y actualiza la UI
        """
        print("[DEBUG - Rutinas] Iniciando load_data")
        try:
            print("[DEBUG - Rutinas] Llamando a routine_controller.get_routines()")
            rutinas = self.routine_controller.get_routines()
            print(f"[DEBUG - Rutinas] routine_controller.get_routines() devolvió {len(rutinas)} rutinas")
            
            if not rutinas:
                print("[DEBUG - Rutinas] No se encontraron rutinas")
                self.show_message("No hay rutinas registradas", ft.colors.ORANGE)
            
            print("[DEBUG - Rutinas] Llamando a update_routines_table")
            self.update_routines_table(rutinas)
            print("[DEBUG - Rutinas] Llamando a page.update()")
            self.page.update()
            print("[DEBUG - Rutinas] load_data finalizado")
            
        except Exception as ex:
            print(f"[DEBUG - Rutinas] Excepción en load_data: {str(ex)}")
            self.show_message(f"Error al cargar las rutinas: {str(ex)}", ft.colors.RED)

    def update_routines_table(self, rutinas):
        """
        Actualiza la tabla con las rutinas
        """
        print(f"[DEBUG - Rutinas] Actualizando tabla con {len(rutinas)} rutinas")
        self.routines_table.rows.clear()
        
        # Actualizar contador de registros
        count = len(rutinas)
        self.records_counter.value = f"{count} {'rutina' if count == 1 else 'rutinas'}"
        
        if not rutinas:
            print("[DEBUG - Rutinas] No hay rutinas para mostrar")
            self.routines_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text("No hay rutinas registradas", italic=True)),
                        ft.DataCell(ft.Text("")),
                        ft.DataCell(ft.Text("")),
                        ft.DataCell(ft.Text("")),
                        ft.DataCell(ft.Text("")),
                    ]
                )
            )
            self.page.update()
            return

        for rutina in rutinas:
            print(f"[DEBUG - Rutinas] Procesando rutina: {rutina.nombre}")
            
            # Contar miembros asignados
            miembros_asignados = self.routine_controller.count_members_assigned_to_routine(rutina.id_rutina)
            
            # Crear botones de acciones
            action_buttons = [
                ft.IconButton(
                    icon=ft.icons.EDIT,
                    icon_color=ft.colors.BLUE,
                    tooltip="Editar",
                    on_click=lambda e, r=rutina: self.edit_routine(r)
                ),
                ft.IconButton(
                    icon=ft.icons.DELETE,
                    icon_color=ft.colors.RED,
                    tooltip="Eliminar",
                    on_click=lambda e, r=rutina: self.delete_routine(r)
                ),
                ft.IconButton(
                    icon=ft.icons.VISIBILITY,
                    icon_color=ft.colors.GREEN,
                    tooltip="Ver detalles",
                    on_click=lambda e, r=rutina: self.view_routine_details(r)
                ),
            ]
            
            self.routines_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(rutina.nombre)),
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(
                                    rutina.nivel_dificultad,
                                    color=ft.colors.WHITE
                                ),
                                bgcolor=ft.colors.BLUE if rutina.nivel_dificultad == "Principiante" else 
                                       ft.colors.ORANGE if rutina.nivel_dificultad == "Intermedio" else 
                                       ft.colors.RED,
                                border_radius=8,
                                padding=ft.padding.symmetric(horizontal=12, vertical=6),
                            )
                        ),
                        ft.DataCell(ft.Text(rutina.descripcion or "-")),
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(
                                    str(miembros_asignados),
                                    text_align=ft.TextAlign.CENTER
                                ),
                                alignment=ft.alignment.center,
                                expand=True
                            )
                        ),
                        ft.DataCell(
                            ft.Row(action_buttons)
                        ),
                    ]
                )
            )
        print("[DEBUG - Rutinas] Tabla actualizada exitosamente")
        self.page.update()

    def show_new_routine_modal(self, e):
        self.file_selected_label = ft.Text("")
        self.selected_file = None
        self.new_routine_modal = ft.AlertDialog(
            title=ft.Text("Nueva Rutina", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.TextField(
                            label="Nombre",
                            width=480,
                            border_radius=8,
                            text_size=16,
                            ref=self.new_routine_name
                        ),
                        ft.Container(height=8),
                        ft.Dropdown(
                            label="Dificultad",
                            width=480,
                            options=[
                                ft.dropdown.Option("Principiante"),
                                ft.dropdown.Option("Intermedio"),
                                ft.dropdown.Option("Avanzado")
                            ],
                            border_radius=8,
                            text_size=16,
                            ref=self.new_routine_difficulty
                        ),
                        ft.Container(height=8),
                        ft.TextField(
                            label="Descripción",
                            width=480,
                            border_radius=8,
                            text_size=16,
                            ref=self.new_routine_description,
                            multiline=True,
                            min_lines=3,
                            max_lines=5
                        ),
                        ft.Container(height=8),
                        ft.ElevatedButton(
                            "Seleccionar Documento de Rutina",
                            icon=ft.icons.UPLOAD_FILE,
                            on_click=lambda _: self.routine_file_picker.pick_files(
                                allow_multiple=False,
                                allowed_extensions=['pdf', 'xlsx', 'xls', 'jpg', 'jpeg', 'png']
                            )
                        ),
                        self.file_selected_label,
                    ],
                    spacing=0,
                ),
                width=540,
                padding=20,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_modal),
                ft.ElevatedButton(
                    "Guardar",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE,
                        color=ft.colors.WHITE,
                    ),
                    on_click=self.save_routine,
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.new_routine_modal not in self.page.overlay:
            self.page.overlay.append(self.new_routine_modal)
        self.new_routine_modal.open = True
        self.page.update()

    def close_modal(self, e):
        """
        Cierra el modal de rutina y limpia los campos
        """
        self.new_routine_modal.open = False
        # Limpiar campos
        self.new_routine_name.current.value = ""
        self.new_routine_difficulty.current.value = None
        self.new_routine_description.current.value = ""
        self.selected_file = None
        self.file_selected_label.value = ""
        self.page.update()

    def save_routine(self, e):
        print("[DEBUG] save_routine llamado")
        try:
            if not self.new_routine_name.current.value:
                print("[DEBUG] Falta nombre")
                self.show_message("El nombre es requerido", ft.colors.RED)
                return
            if not self.new_routine_difficulty.current.value:
                print("[DEBUG] Falta dificultad")
                self.show_message("La dificultad es requerida", ft.colors.RED)
                return
            if not hasattr(self, 'selected_file'):
                print("[DEBUG] Falta archivo")
                self.show_message("Por favor, seleccione un archivo de rutina", ft.colors.RED)
                return
            
            # Validar tamaño del archivo (máximo 1MB para compatibilidad)
            MAX_FILE_SIZE = 1 * 1024 * 1024  # 1MB en bytes
            file_size = os.path.getsize(self.selected_file.path)
            if file_size > MAX_FILE_SIZE:
                self.show_message(f"El archivo es demasiado grande. Máximo permitido: 1MB. Tu archivo: {file_size / (1024*1024):.1f}MB", ft.colors.RED)
                return
            
            print("[DEBUG] Todos los campos requeridos presentes")
            with open(self.selected_file.path, 'rb') as file:
                file_content = file.read()
            import datetime
            now = datetime.datetime.now()
            fecha_creacion = now
            fecha_horario = now
            routine_data = {
                'nombre': self.new_routine_name.current.value,
                'nivel_dificultad': self.new_routine_difficulty.current.value,
                'descripcion': self.new_routine_description.current.value,
                'documento_rutina': file_content,
                'fecha_creacion': fecha_creacion,
                'fecha_horario': fecha_horario,
            }
            # Refuerzo: si por alguna razón no están, los agrego
            if not routine_data.get('fecha_creacion'):
                routine_data['fecha_creacion'] = now
            if not routine_data.get('fecha_horario'):
                routine_data['fecha_horario'] = now
            print(f"[DEBUG] routine_data a guardar: {routine_data}")
            success, message = self.routine_controller.create_routine(routine_data)
            print(f"[DEBUG] Resultado create_routine: success={success}, message={message}")
            if success:
                self.close_modal(e)
                self.load_data()
                self.show_message("Rutina creada exitosamente", ft.colors.GREEN)
            else:
                self.show_message(f"Error al crear la rutina: {message}", ft.colors.RED)
        except Exception as e:
            print(f"[DEBUG] Excepción inesperada en save_routine: {str(e)}")
            self.show_message(f"Error inesperado: {str(e)}", ft.colors.RED)

    def edit_routine(self, rutina):
        self.editing_routine = rutina
        self.file_selected_label = ft.Text("")
        self.selected_file = None
        self.new_routine_modal = ft.AlertDialog(
            title=ft.Text("Editar Rutina", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.TextField(
                            label="Nombre",
                            width=480,
                            border_radius=8,
                            text_size=16,
                            ref=self.new_routine_name,
                            value=rutina.nombre
                        ),
                        ft.Container(height=8),
                        ft.Dropdown(
                            label="Dificultad",
                            width=480,
                            options=[
                                ft.dropdown.Option("Principiante"),
                                ft.dropdown.Option("Intermedio"),
                                ft.dropdown.Option("Avanzado")
                            ],
                            border_radius=8,
                            text_size=16,
                            ref=self.new_routine_difficulty,
                            value=rutina.nivel_dificultad
                        ),
                        ft.Container(height=8),
                        ft.TextField(
                            label="Descripción",
                            width=480,
                            border_radius=8,
                            text_size=16,
                            ref=self.new_routine_description,
                            multiline=True,
                            min_lines=3,
                            max_lines=5,
                            value=rutina.descripcion
                        ),
                        ft.Container(height=8),
                        ft.ElevatedButton(
                            "Seleccionar Nuevo Documento",
                            icon=ft.icons.UPLOAD_FILE,
                            on_click=lambda _: self.routine_file_picker.pick_files(
                                allow_multiple=False,
                                allowed_extensions=['pdf', 'xlsx', 'xls', 'jpg', 'jpeg', 'png']
                            )
                        ),
                        self.file_selected_label,
                    ],
                    spacing=0,
                ),
                width=540,
                padding=20,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_modal),
                ft.ElevatedButton(
                    "Actualizar",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE,
                        color=ft.colors.WHITE,
                    ),
                    on_click=self.update_routine,
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.new_routine_modal not in self.page.overlay:
            self.page.overlay.append(self.new_routine_modal)
        self.new_routine_modal.open = True
        self.page.update()

    def update_routine(self, e):
        try:
            if not self.new_routine_name.current.value:
                self.show_message("El nombre es requerido", ft.colors.RED)
                return
            if not self.new_routine_difficulty.current.value:
                self.show_message("La dificultad es requerida", ft.colors.RED)
                return
            routine_data = {
                'nombre': self.new_routine_name.current.value,
                'nivel_dificultad': self.new_routine_difficulty.current.value,
                'descripcion': self.new_routine_description.current.value,
            }
            if hasattr(self, 'selected_file') and self.selected_file:
                # Validar tamaño del archivo (máximo 1MB para compatibilidad)
                MAX_FILE_SIZE = 1 * 1024 * 1024  # 1MB en bytes
                file_size = os.path.getsize(self.selected_file.path)
                if file_size > MAX_FILE_SIZE:
                    self.show_message(f"El archivo es demasiado grande. Máximo permitido: 1MB. Tu archivo: {file_size / (1024*1024):.1f}MB", ft.colors.RED)
                    return
                
                with open(self.selected_file.path, 'rb') as file:
                    routine_data['documento_rutina'] = file.read()
            success, message = self.routine_controller.update_routine(self.editing_routine.id_rutina, routine_data)
            if success:
                self.close_modal(e)
                self.load_data()
                self.show_message("Rutina actualizada exitosamente", ft.colors.GREEN)
            else:
                self.show_message(f"Error al actualizar la rutina: {message}", ft.colors.RED)
        except Exception as e:
            self.show_message(f"Error inesperado: {str(e)}", ft.colors.RED)

    def delete_routine(self, rutina):
        self.routine_to_delete = rutina
        self.confirm_dialog.open = True
        self.page.update()

    def _close_dialog(self, dialog):
        dialog.open = False
        self.page.update()

    def view_routine_details(self, rutina):
        controls = [
            ft.Text("Información General", size=16, weight=ft.FontWeight.BOLD),
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text("Nombre:", size=14, weight=ft.FontWeight.BOLD),
                        ft.Text(rutina.nombre, size=14),
                    ]),
                    ft.Row([
                        ft.Text("Dificultad:", size=14, weight=ft.FontWeight.BOLD),
                        ft.Text(rutina.nivel_dificultad, size=14),
                    ]),
                    ft.Row([
                        ft.Text("Descripción:", size=14, weight=ft.FontWeight.BOLD),
                        ft.Text(rutina.descripcion or "-", size=14),
                    ]),
                    ft.Row([
                        ft.Text("Fecha de Creación:", size=14, weight=ft.FontWeight.BOLD),
                        ft.Text(rutina.fecha_creacion.strftime("%d/%m/%Y %H:%M") if rutina.fecha_creacion else "-", size=14),
                    ]),
                ], spacing=10),
                padding=10,
            ),
            ft.Divider(),
        ]
        # Previsualización o descarga de archivo
        if hasattr(rutina, 'documento_rutina') and rutina.documento_rutina:
            controls.append(ft.Text("Documento de Rutina", size=16, weight=ft.FontWeight.BOLD))
            controls.append(
                ft.ElevatedButton(
                    "Previsualizar/Descargar Archivo",
                    icon=ft.icons.PICTURE_AS_PDF,
                    on_click=lambda e, r=rutina: self._preview_routine_file(r)
                )
            )
        details_modal = ft.AlertDialog(
            title=ft.Text(f"Detalles de la Rutina", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                content=ft.Column(controls=controls, spacing=0),
                width=600,
                padding=20,
            ),
            actions=[
                ft.TextButton("Cerrar", on_click=lambda e: self._close_dialog(details_modal)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if details_modal not in self.page.overlay:
            self.page.overlay.append(details_modal)
        details_modal.open = True
        self.page.update()

    def _preview_routine_file(self, rutina):
        import tempfile, os, webbrowser
        # Intentar determinar extensión (por ahora PDF por defecto)
        ext = '.pdf'
        # Si quieres mejorar esto, puedes guardar el nombre original del archivo y extraer la extensión
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
            tmp_file.write(rutina.documento_rutina)
            tmp_file.flush()
            webbrowser.open(tmp_file.name)

    def apply_filters(self, e):
        """
        Aplica los filtros seleccionados
        """
        try:
            filters = {}
            
            if self.search_field.value:
                filters['search'] = self.search_field.value
            
            if self.difficulty_filter.value and self.difficulty_filter.value != "Todas":
                filters['nivel_dificultad'] = self.difficulty_filter.value
            
            print(f"[DEBUG] Aplicando filtros: {filters}")  # Debug log
            rutinas = self.routine_controller.get_routines(filters)
            print(f"[DEBUG] Rutinas encontradas: {len(rutinas)}")  # Debug log
            self.update_routines_table(rutinas)
        except Exception as ex:
            print(f"[DEBUG] Error al aplicar filtros: {str(ex)}")  # Debug log
            self.show_message(f"Error al aplicar filtros: {str(ex)}", ft.colors.RED)

    def clear_filters(self, e):
        self.search_field.value = ""
        self.difficulty_filter.value = "Todas"
        self.page.update()
        self.load_data()

    def export_to_excel(self, e):
        """
        Exporta las rutinas actuales a un archivo Excel
        """
        try:
            self.export_type = "excel"  # Establecer tipo de exportación
            
            # Obtener las rutinas filtradas actuales
            filters = {}
            
            if self.search_field.value:
                filters['search'] = self.search_field.value
            
            if self.difficulty_filter.value and self.difficulty_filter.value != "Todas":
                filters['nivel_dificultad'] = self.difficulty_filter.value

            rutinas = self.routine_controller.get_routines(filters)
            
            if not rutinas:
                self.show_message("No hay rutinas para exportar", ft.colors.ORANGE)
                return

            # Actualizar el contenido del diálogo con el número de rutinas
            self.export_dialog.content.value = f"¿Deseas exportar {len(rutinas)} rutinas a Excel?\nEl archivo se guardará en tu carpeta de Descargas."
            
            # Guardar las rutinas en una variable de instancia para usarla en confirm_export
            self.rutinas_to_export = rutinas
            
            # Mostrar el diálogo
            self.export_dialog.open = True
            self.page.update()

        except Exception as e:
            self.show_message(f"Error al preparar la exportación: {str(e)}", ft.colors.RED)

    def export_to_pdf(self, e):
        """
        Exporta las rutinas actuales a un archivo PDF
        """
        try:
            self.export_type = "pdf"  # Establecer tipo de exportación
            
            # Obtener las rutinas filtradas actuales
            filters = {}
            
            if self.search_field.value:
                filters['search'] = self.search_field.value
            
            if self.difficulty_filter.value and self.difficulty_filter.value != "Todas":
                filters['nivel_dificultad'] = self.difficulty_filter.value

            rutinas = self.routine_controller.get_routines(filters)
            
            if not rutinas:
                self.show_message("No hay rutinas para exportar", ft.colors.ORANGE)
                return

            # Actualizar el contenido del diálogo con el número de rutinas
            self.export_dialog.content.value = f"¿Deseas exportar {len(rutinas)} rutinas a PDF?\nEl archivo se guardará en tu carpeta de Descargas."
            
            # Guardar las rutinas en una variable de instancia para usarla en confirm_export
            self.rutinas_to_export = rutinas
            
            # Mostrar el diálogo
            self.export_dialog.open = True
            self.page.update()

        except Exception as e:
            self.show_message(f"Error al preparar la exportación: {str(e)}", ft.colors.RED)

    def confirm_export(self, e):
        """
        Confirma la exportación y genera el archivo
        """
        try:
            # Obtener la ruta de la carpeta de descargas
            downloads_path = str(Path.home() / "Downloads")

            # Usar las rutinas guardadas en export_to_excel/export_to_pdf
            rutinas = getattr(self, 'rutinas_to_export', [])
            if not rutinas:
                self.show_message("No hay rutinas para exportar", ft.colors.ORANGE)
                return

            # Determinar el tipo de exportación
            if self.export_type == "excel":
                self._export_to_excel(rutinas, downloads_path)
            elif self.export_type == "pdf":
                self._export_to_pdf(rutinas, downloads_path)
            else:
                self.show_message("Error: Tipo de exportación no válido", ft.colors.RED)
                return

            # Limpiar las rutinas guardadas
            self.rutinas_to_export = None

            # Cerrar el diálogo de confirmación
            self.close_export_dialog(e)

        except Exception as e:
            self.show_message(f"Error al exportar: {str(e)}", ft.colors.RED)

    def close_export_dialog(self, e):
        """
        Cierra el diálogo de exportación
        """
        self.export_dialog.open = False
        self.export_type = None  # Resetear el tipo de exportación
        self.page.update()

    def _export_to_excel(self, rutinas, downloads_path):
        """
        Exporta las rutinas a Excel
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rutinas"

        # Estilos
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        date_format = 'dd/mm/yyyy'

        # Configurar ancho de columnas
        ws.column_dimensions['A'].width = 30  # Nombre
        ws.column_dimensions['B'].width = 20  # Dificultad
        ws.column_dimensions['C'].width = 40  # Descripción
        ws.column_dimensions['D'].width = 20  # Miembros Asignados
        ws.column_dimensions['E'].width = 20  # Fecha Creación

        # Escribir encabezados
        headers = ["Nombre", "Dificultad", "Descripción", "Miembros Asignados", "Fecha Creación"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Escribir datos
        for row, rutina in enumerate(rutinas, 2):
            # Nombre
            cell = ws.cell(row=row, column=1)
            cell.value = rutina.nombre
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Dificultad
            cell = ws.cell(row=row, column=2)
            cell.value = rutina.nivel_dificultad
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color de fondo según dificultad
            if rutina.nivel_dificultad == "Principiante":
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            elif rutina.nivel_dificultad == "Intermedio":
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            elif rutina.nivel_dificultad == "Avanzado":
                cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

            # Descripción
            cell = ws.cell(row=row, column=3)
            cell.value = rutina.descripcion if rutina.descripcion else ""
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Miembros Asignados
            cell = ws.cell(row=row, column=4)
            miembros_asignados = self.routine_controller.count_members_assigned_to_routine(rutina.id_rutina)
            cell.value = miembros_asignados
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Fecha Creación
            cell = ws.cell(row=row, column=5)
            if rutina.fecha_creacion:
                cell.value = rutina.fecha_creacion
                cell.number_format = date_format
            else:
                cell.value = ""
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Congelar la primera fila
        ws.freeze_panes = 'A2'

        # Generar nombre del archivo
        filename = f"rutinas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(downloads_path, filename)

        # Guardar archivo
        wb.save(filepath)

        # Mostrar mensaje de éxito
        self.show_message(f"Archivo Excel guardado en: {filepath}", ft.colors.GREEN)

    def _export_to_pdf(self, rutinas, downloads_path):
        """
        Exporta las rutinas a PDF
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from datetime import datetime

            # Generar nombre del archivo
            filename = f"rutinas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(downloads_path, filename)

            # Crear el documento PDF
            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(letter),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )

            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Centrado
            )

            # Contenido del PDF
            elements = []

            # Título
            title = Paragraph("Reporte de Rutinas", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))

            # Información de filtros aplicados
            filter_info = []
            if self.search_field.value:
                filter_info.append(f"Búsqueda: {self.search_field.value}")
            if self.difficulty_filter.value and self.difficulty_filter.value != "Todas":
                filter_info.append(f"Dificultad: {self.difficulty_filter.value}")

            if filter_info:
                filter_text = " | ".join(filter_info)
                filter_paragraph = Paragraph(f"Filtros aplicados: {filter_text}", styles["Normal"])
                elements.append(filter_paragraph)
                elements.append(Spacer(1, 20))

            # Datos de la tabla
            data = [["Nombre", "Dificultad", "Descripción", "Miembros Asignados", "Fecha Creación"]]
            
            for rutina in rutinas:
                miembros_asignados = self.routine_controller.count_members_assigned_to_routine(rutina.id_rutina)
                data.append([
                    rutina.nombre,
                    rutina.nivel_dificultad,
                    rutina.descripcion if rutina.descripcion else "",
                    str(miembros_asignados),
                    rutina.fecha_creacion.strftime("%d/%m/%Y") if rutina.fecha_creacion else ""
                ])

            # Crear la tabla
            table = Table(data, colWidths=[2.5*inch, 1.5*inch, 3*inch, 1.5*inch, 1.5*inch])
            
            # Estilo de la tabla
            table_style = TableStyle([
                # Encabezados
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Alineación específica
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Nombre
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Descripción
            ])

            # Agregar colores de fondo según dificultad
            for i, rutina in enumerate(rutinas, 1):
                if rutina.nivel_dificultad == "Principiante":
                    table_style.add('BACKGROUND', (1, i), (1, i), colors.HexColor('#E3F2FD'))
                elif rutina.nivel_dificultad == "Intermedio":
                    table_style.add('BACKGROUND', (1, i), (1, i), colors.HexColor('#FFF3E0'))
                elif rutina.nivel_dificultad == "Avanzado":
                    table_style.add('BACKGROUND', (1, i), (1, i), colors.HexColor('#FFEBEE'))

            table.setStyle(table_style)
            elements.append(table)

            # Pie de página con fecha y hora
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=1
            )
            footer = Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                footer_style
            )
            elements.append(Spacer(1, 20))
            elements.append(footer)

            # Generar el PDF
            doc.build(elements)

            # Mostrar mensaje de éxito
            self.show_message(f"Archivo PDF guardado en: {filepath}", ft.colors.GREEN)

        except Exception as e:
            self.show_message(f"Error al generar PDF: {str(e)}", ft.colors.RED)

    def show_message(self, content, bgcolor: str):
        """
        Muestra un mensaje temporal
        """
        self.page.snack_bar = ft.SnackBar(
            content=ft.Text(content),
            bgcolor=bgcolor,
            open=True,
        )
        self.page.update() 