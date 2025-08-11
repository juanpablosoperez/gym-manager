import flet as ft
import logging
from gym_manager.views.module_views import ModuleView
from gym_manager.models.payment import Pago
from gym_manager.controllers.payment_controller import PaymentController
from gym_manager.utils.database import get_db_session
from gym_manager.controllers.monthly_fee_controller import MonthlyFeeController
from datetime import datetime, timedelta
from gym_manager.utils.navigation import navigate_to_login
from gym_manager.models.member import Miembro
from gym_manager.models.payment_method import MetodoPago
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from pathlib import Path
from gym_manager.models.monthly_fee import CuotaMensual
from gym_manager.utils.database import session_scope
from sqlalchemy.orm import joinedload
import subprocess
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Configurar logger
logger = logging.getLogger(__name__)

class PaymentsView(ModuleView):
    def __init__(self, page: ft.Page):
        # Inicializar variables básicas
        self.page = page
        self.title = "Gestión de Pagos"
        self.content = None
        
        # Inicializar controladores
        self.payment_controller = PaymentController(get_db_session())
        self.monthly_fee_controller = MonthlyFeeController(get_db_session())
        self.db_session = get_db_session()
        self.current_monthly_fee = None  # Variable para almacenar la cuota mensual actual
        
        # Inicializar paginación ANTES de setup_payment_view
        from gym_manager.utils.pagination import PaginationController, PaginationWidget
        self.pagination_controller = PaginationController(items_per_page=10)
        self.pagination_widget = PaginationWidget(
            self.pagination_controller, 
            on_page_change=self._on_page_change
        )
        
        # Ahora llamar setup_payment_view después de inicializar todo
        self.setup_payment_view()
        self.export_type = None  # Agregar variable para el tipo de exportación
        self.check_overdue_payments()  # Verificar pagos vencidos al iniciar
        # NO llamar load_data aquí, se llamará cuando la vista se muestre

    def setup_payment_view(self):
        # Título amigable arriba de los filtros, en negro y bien arriba
        self.title = ft.Text(
            "Gestión de Pagos",
            size=32,
            weight=ft.FontWeight.BOLD,
            color=ft.colors.BLACK,
        )

        # Botón Nuevo Pago
        self.new_payment_btn = ft.ElevatedButton(
            text="Nuevo Pago",
            icon=ft.icons.ADD,
            style=ft.ButtonStyle(
                bgcolor=ft.colors.BLUE,
                color=ft.colors.WHITE,
                shape=ft.RoundedRectangleBorder(radius=10),
                padding=ft.padding.symmetric(horizontal=10, vertical=8),
                text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD),
            ),
            width=150,
            on_click=self.show_new_payment_modal
        )

        # Botón Exportar a Excel (solo icono y tooltip)
        self.export_excel_btn = ft.IconButton(
            icon=ft.icons.TABLE_VIEW,
            icon_color=ft.colors.GREEN_700,
            tooltip="Exportar a Excel",
            on_click=self.export_to_excel,
            width=48,
            height=48,
        )
        # Botón Exportar a PDF (solo icono y tooltip)
        self.export_pdf_btn = ft.IconButton(
            icon=ft.icons.PICTURE_AS_PDF,
            icon_color=ft.colors.RED_700,
            tooltip="Exportar a PDF",
            on_click=self.export_to_pdf,
            width=48,
            height=48,
        )

        # Filtros de búsqueda
        self.search_field = ft.TextField(
            label="Buscar por nombre",
            prefix_icon=ft.icons.SEARCH,
            border_radius=10,
            width=260,
            height=48,
            text_size=16,
            on_change=self.apply_filters
        )

        # DatePicker Desde
        self.date_from = ft.DatePicker(
            first_date=datetime(2024, 1, 1),
            last_date=datetime(2025, 12, 31),
            on_change=self.on_date_from_change
        )
        self.page.overlay.append(self.date_from)
        self.date_from_value = None
        self.date_from_field = ft.Container(
            content=ft.Row([
                ft.Text(self.date_from_value.strftime("%d/%m/%Y") if self.date_from_value else "",
                        size=16,
                        color=ft.colors.BLACK54),
                ft.Icon(ft.icons.CALENDAR_TODAY, size=20, color=ft.colors.GREY_700),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            border=ft.border.all(1, ft.colors.GREY_400),
            border_radius=10,
            padding=ft.padding.symmetric(horizontal=12, vertical=8),
            width=120,
            on_click=lambda e: self.open_date_picker(self.date_from),
            bgcolor=ft.colors.WHITE,
            tooltip="Seleccionar fecha desde",
        )

        # DatePicker Hasta
        self.date_to = ft.DatePicker(
            first_date=datetime(2024, 1, 1),
            last_date=datetime(2025, 12, 31),
            on_change=self.on_date_to_change
        )
        self.page.overlay.append(self.date_to)
        self.date_to_value = None
        self.date_to_field = ft.Container(
            content=ft.Row([
                ft.Text(self.date_to_value.strftime("%d/%m/%Y") if self.date_to_value else "", 
                        size=16, 
                        color=ft.colors.BLACK54),
                ft.Icon(ft.icons.CALENDAR_TODAY, size=20, color=ft.colors.GREY_700),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            border=ft.border.all(1, ft.colors.GREY_400),
            border_radius=10,
            padding=ft.padding.symmetric(horizontal=12, vertical=8),
            width=120,
            on_click=lambda e: self.open_date_picker(self.date_to),
            bgcolor=ft.colors.WHITE,
            tooltip="Seleccionar fecha hasta",
        )

        self.payment_method = ft.Dropdown(
            label="Método de pago",
            width=180,
            options=[
                ft.dropdown.Option("Todos"),
                ft.dropdown.Option("Efectivo"),
                ft.dropdown.Option("Transferencia bancaria"),
                ft.dropdown.Option("Tarjeta de crédito")
            ],
            border_radius=10,
            text_size=16,
            on_change=self.apply_filters
        )

        # Filtro de estado
        self.status_filter = ft.Dropdown(
            label="Estado",
            width=120,
            options=[
                ft.dropdown.Option("Todos"),
                ft.dropdown.Option("Pagado"),
                ft.dropdown.Option("Cancelado")
            ],
            border_radius=10,
            text_size=16,
            value="Pagado",
            on_change=self.apply_filters
        )

        self.clear_btn = ft.OutlinedButton(
            text="Limpiar filtros",
            icon=ft.icons.CLEAR,
            style=ft.ButtonStyle(
                color=ft.colors.GREY_700,
                shape=ft.RoundedRectangleBorder(radius=8),
                padding=ft.padding.symmetric(horizontal=18, vertical=12),
                text_style=ft.TextStyle(size=16),
            )
        )

        self.clear_btn.on_click = self.clear_filters

        # Tabla de pagos
        self.payments_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Miembro", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Fecha", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Monto", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Método", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Observaciones", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Estado", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Acciones", size=18, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=12,
            vertical_lines=ft.border.all(1, ft.colors.GREY_300),
            horizontal_lines=ft.border.all(1, ft.colors.GREY_300),
            column_spacing=60,
            heading_row_color=ft.colors.GREY_100,
            heading_row_height=60,
            data_row_color=ft.colors.WHITE,
            data_row_min_height=56,
        )

        # Modal de nuevo pago
        self.new_payment_client_field = ft.TextField(
            label="Miembro",
            prefix_icon=ft.icons.SEARCH,
            border_radius=8,
            hint_text="Buscar miembro por nombre o documento...",
            width=500,
            height= 40,
            on_change=self.search_member
        )
        
        # Primero define el ListView
        self.member_search_results = ft.ListView(
            spacing=10,
            padding=10,
            height=40,
            visible=False,
            auto_scroll=False,
            expand=False
        )
        # Luego el Container que lo envuelve
        self.member_search_results_container = ft.Container(
            content=self.member_search_results,
            height=0 if not self.member_search_results.visible else 60,  # altura dinámica
            padding=0,
            margin=0
        )
        
        self.selected_member = None
        self.new_payment_date_picker = ft.DatePicker(
            first_date=datetime(2024, 1, 1),
            last_date=datetime(2025, 12, 31),
            on_change=self.on_new_payment_date_change
        )
        self.page.overlay.append(self.new_payment_date_picker)
        self.new_payment_date_value = None
        self.new_payment_date_field = ft.Container(
            content=ft.Row([
                ft.Text(self.new_payment_date_value.strftime("%d/%m/%Y") if self.new_payment_date_value else "Seleccionar fecha",
                        size=16,
                        color=ft.colors.BLACK54),
                ft.Icon(ft.icons.CALENDAR_TODAY, size=20, color=ft.colors.GREY_700),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=8,
            padding=ft.padding.symmetric(horizontal=12, vertical=10),
            width=500,
            on_click=lambda e: self.open_date_picker(self.new_payment_date_picker),
            bgcolor=ft.colors.GREY_100,
            tooltip="Seleccionar fecha",
        )
        self.new_payment_amount_field = ft.TextField(
            label="Monto",
            prefix_icon=ft.icons.ATTACH_MONEY,
            border_radius=8,
            width=500,
            value="0.00",
            on_change=self.validate_payment_amount
        )
        
        # Mensaje de advertencia para el monto
        self.amount_warning_text = ft.Text(
            "",
            color=ft.colors.RED,
            size=12,
            visible=False
        )

        # Actualizar el dropdown de métodos de pago para que cargue los métodos activos
        self.new_payment_method_field = ft.Dropdown(
            label="Método de pago",
            options=[],  # Se llenará dinámicamente
            border_radius=8,
            width=500,
            hint_text="Seleccionar método",
        )
        self.new_payment_observations_field = ft.TextField(
            label="Observaciones",
            multiline=True,
            min_lines=3,
            max_lines=5,
            border_radius=8,
            width=500,
            height=60,
            hint_text="Agregar observaciones...",
        )
        # Mensaje de validación para el comprobante
        self.receipt_validation_text = ft.Text(
            "",
            color=ft.colors.GREEN,
            size=12,
            visible=False
        )

        # Modal de nuevo pago
        self.new_payment_modal = ft.AlertDialog(
            title=ft.Text("Nuevo Pago", size=22, weight=ft.FontWeight.BOLD),
            content=ft.Column(
                controls=[
                    ft.Text("Miembro", size=16, weight=ft.FontWeight.BOLD),
                    self.new_payment_client_field,
                    self.member_search_results_container,
                    ft.Text("Fecha", size=16, weight=ft.FontWeight.BOLD),
                    self.new_payment_date_field,
                    ft.Text("Monto", size=16, weight=ft.FontWeight.BOLD),
                    self.new_payment_amount_field,
                    self.amount_warning_text,  # Agregar el mensaje de advertencia aquí
                    ft.Text("Método de pago", size=16, weight=ft.FontWeight.BOLD),
                    self.new_payment_method_field,
                    ft.Text("Observaciones", size=16, weight=ft.FontWeight.BOLD),
                    self.new_payment_observations_field,
                    self.receipt_validation_text,  # Agregar el texto de validación
                    ft.Row(
                        controls=[
                            ft.TextButton("Cancelar", on_click=self.close_modal, style=ft.ButtonStyle(bgcolor=ft.colors.WHITE, color=ft.colors.BLACK87, shape=ft.RoundedRectangleBorder(radius=8), padding=ft.padding.symmetric(horizontal=28, vertical=12), text_style=ft.TextStyle(size=18, weight=ft.FontWeight.BOLD))),
                            ft.ElevatedButton(
                                "Comprobante",
                                icon=ft.icons.RECEIPT,
                                style=ft.ButtonStyle(
                                    bgcolor=ft.colors.GREEN_700,
                                    color=ft.colors.WHITE,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                    padding=ft.padding.symmetric(horizontal=20, vertical=12),
                                    text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD),
                                ),
                                on_click=self.generate_receipt_from_form
                            ),
                            ft.ElevatedButton(
                                "Guardar",
                                style=ft.ButtonStyle(
                                    bgcolor=ft.colors.BLUE_900,
                                    color=ft.colors.WHITE,
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                    padding=ft.padding.symmetric(horizontal=28, vertical=12),
                                    text_style=ft.TextStyle(size=18, weight=ft.FontWeight.BOLD),
                                ),
                                on_click=self.save_payment
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.END,
                    ),
                ],
                spacing=20,
            ),
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True,
        )

        # Modal de edición de pago
        self.edit_payment_client_field = ft.TextField(
            label="Miembro",
            prefix_icon=ft.icons.PERSON,
            border_radius=8,
            width=500,
            read_only=True
        )
        self.edit_payment_date_picker = ft.DatePicker(
            first_date=datetime(2024, 1, 1),
            last_date=datetime(2025, 12, 31),
            on_change=self.on_edit_payment_date_change
        )
        self.page.overlay.append(self.edit_payment_date_picker)
        self.edit_payment_date_value = None
        self.edit_payment_date_field = ft.Container(
            content=ft.Row([
                ft.Text(self.edit_payment_date_value.strftime("%d/%m/%Y") if self.edit_payment_date_value else "Seleccionar fecha",
                        size=16,
                        color=ft.colors.BLACK54),
                ft.Icon(ft.icons.CALENDAR_TODAY, size=20, color=ft.colors.GREY_700),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=8,
            padding=ft.padding.symmetric(horizontal=12, vertical=10),
            width=500,
            on_click=lambda e: self.open_date_picker(self.edit_payment_date_picker),
            bgcolor=ft.colors.GREY_100,
            tooltip="Seleccionar fecha",
        )
        self.edit_payment_amount_field = ft.TextField(
            label="Monto",
            prefix_icon=ft.icons.ATTACH_MONEY,
            border_radius=8,
            width=500,
            value="0.00",
        )
        self.edit_payment_method_field = ft.Dropdown(
            label="Método de pago",
            options=[],  # Se llenará dinámicamente
            border_radius=8,
            width=500,
            hint_text="Seleccionar método",
        )
        self.edit_payment_observations_field = ft.TextField(
            label="Observaciones",
            multiline=True,
            min_lines=3,
            max_lines=5,
            border_radius=8,
            width=500,
            height=60,
            hint_text="Agregar observaciones...",
        )
        self.edit_payment_modal = ft.AlertDialog(
            title=ft.Text("Editar Pago", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Column(
                controls=[
                    ft.Text("Miembro", size=16, weight=ft.FontWeight.BOLD),
                    self.edit_payment_client_field,
                    ft.Text("Fecha", size=16, weight=ft.FontWeight.BOLD),
                    self.edit_payment_date_field,
                    ft.Text("Monto", size=16, weight=ft.FontWeight.BOLD),
                    self.edit_payment_amount_field,
                    ft.Text("Método de pago", size=16, weight=ft.FontWeight.BOLD),
                    self.edit_payment_method_field,
                    ft.Text("Observaciones", size=16, weight=ft.FontWeight.BOLD),
                    self.edit_payment_observations_field,
                ],
                spacing=18,
                width=540,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_edit_modal, style=ft.ButtonStyle(bgcolor=ft.colors.WHITE, color=ft.colors.BLACK87, shape=ft.RoundedRectangleBorder(radius=8), padding=ft.padding.symmetric(horizontal=28, vertical=12), text_style=ft.TextStyle(size=18, weight=ft.FontWeight.BOLD))),
                ft.ElevatedButton(
                    "Actualizar",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE_900,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=18, weight=ft.FontWeight.BOLD),
                    ),
                    on_click=self.update_payment
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True,
        )
        self.page.overlay.append(self.edit_payment_modal)

        # Modal de confirmación de borrado
        self.delete_confirm_modal = ft.AlertDialog(
            title=ft.Text("Confirmar Cancelación", size=22, weight=ft.FontWeight.BOLD),
            content=ft.Text("¿Estás seguro que deseas cancelar este pago? El pago cancelado no se mostrará más en la lista.", size=16),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_delete_modal, style=ft.ButtonStyle(bgcolor=ft.colors.WHITE, color=ft.colors.BLACK87, shape=ft.RoundedRectangleBorder(radius=8), padding=ft.padding.symmetric(horizontal=28, vertical=12), text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD))),
                ft.ElevatedButton(
                    "Confirmar Cancelación",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.RED_700,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD),
                    ),
                    on_click=self.confirm_delete_payment
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True,
        )
        self.page.overlay.append(self.delete_confirm_modal)
        self.selected_payment_to_delete = None

        # Modal de confirmación de comprobante
        self.receipt_confirm_modal = ft.AlertDialog(
            title=ft.Text("Generar Comprobante", size=22, weight=ft.FontWeight.BOLD),
            content=ft.Text(
                "¿Deseas generar el comprobante de pago?\n"
                "El archivo se guardará en tu carpeta de Descargas.",
                size=16
            ),
            actions=[
                ft.TextButton(
                    "Cancelar",
                    on_click=self.close_receipt_modal,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.WHITE,
                        color=ft.colors.BLACK87,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
                ft.ElevatedButton(
                    "Generar",
                    on_click=self.confirm_generate_receipt,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE_900,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True,
        )
        self.page.overlay.append(self.receipt_confirm_modal)
        self.selected_payment_for_receipt = None

        # Agregar el indicador de cuota mensual
        self.monthly_fee_card = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Row(
                        controls=[
                            ft.Icon(
                                name=ft.icons.ATTACH_MONEY,
                                color=ft.colors.BLUE_900,
                                size=32
                            ),
                            ft.Column(
                                controls=[
                                    ft.Text(
                                        "Cuota Mensual",
                                        size=16,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.colors.BLUE_900
                                    ),
                                    ft.Text(
                                        f"${self.monthly_fee_controller.get_current_fee().monto:,.2f}" if self.monthly_fee_controller.get_current_fee() else "No configurada",
                                        size=24,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.colors.BLUE_900
                                    )
                                ],
                                spacing=2
                            ),
                            ft.IconButton(
                                icon=ft.icons.EDIT,
                                icon_color=ft.colors.BLUE_900,
                                tooltip="Editar cuota mensual",
                                on_click=self.show_edit_fee_modal
                            )
                        ],
                        alignment=ft.MainAxisAlignment.START,
                        spacing=10
                    )
                ],
                spacing=5
            ),
            padding=20,
            border_radius=12,
            bgcolor=ft.colors.WHITE,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=10,
                color=ft.colors.GREY_300,
            )
        )

        # Modal para editar la cuota mensual
        self.edit_fee_field = ft.TextField(
            label="Nueva cuota mensual",
            prefix_icon=ft.icons.ATTACH_MONEY,
            border_radius=8,
            width=300,
            height=200,
            value=str(self.monthly_fee_controller.get_current_fee().monto) if self.monthly_fee_controller.get_current_fee() else "0.00"
        )

        # DatePicker para la fecha de actualización
        self.edit_fee_date_picker = ft.DatePicker(
            first_date=datetime(2024, 1, 1),
            last_date=datetime(2025, 12, 31),
            on_change=self.on_edit_fee_date_change
        )
        self.page.overlay.append(self.edit_fee_date_picker)
        self.edit_fee_date_value = None
        self.edit_fee_date_field = ft.Container(
            content=ft.Row([
                ft.Text(self.edit_fee_date_value.strftime("%d/%m/%Y") if self.edit_fee_date_value else "Seleccionar fecha",
                        size=16,
                        color=ft.colors.BLACK54),
                ft.Icon(ft.icons.CALENDAR_TODAY, size=20, color=ft.colors.GREY_700),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=8,
            padding=ft.padding.symmetric(horizontal=12, vertical=10),
            width=300,
            on_click=lambda e: self.open_date_picker(self.edit_fee_date_picker),
            bgcolor=ft.colors.GREY_100,
            tooltip="Seleccionar fecha",
        )

        self.edit_fee_modal = ft.AlertDialog(
            title=ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Icon(
                            name=ft.icons.ATTACH_MONEY,
                            color=ft.colors.BLUE_900,
                            size=32
                        ),
                        ft.Text(
                            "Editar Cuota Mensual",
                            size=24,
                            weight=ft.FontWeight.BOLD,
                            color=ft.colors.BLUE_900
                        )
                    ],
                    alignment=ft.MainAxisAlignment.CENTER,
                    spacing=10
                ),
                padding=ft.padding.only(bottom=20)
            ),
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Container(
                            content=ft.Column(
                                controls=[
                                    ft.Text(
                                        "Monto Actual",
                                        size=14,
                                        color=ft.colors.GREY_700
                                    ),
                                    ft.Text(
                                        f"${self.monthly_fee_controller.get_current_fee().monto:,.2f}" if self.monthly_fee_controller.get_current_fee() else "$0.00",
                                        size=24,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.colors.BLUE_900
                                    ),
                                    ft.Text(
                                        f"Última actualización: {self.monthly_fee_controller.get_current_fee().fecha_actualizacion.strftime('%d/%m/%Y')}" if self.monthly_fee_controller.get_current_fee() else "No hay fecha de actualización",
                                        size=12,
                                        color=ft.colors.GREY_600
                                    )
                                ],
                                spacing=2,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                            padding=ft.padding.only(bottom=20)
                        ),
                        ft.Container(
                            content=ft.Divider(height=1, color=ft.colors.GREY_300),
                            margin=ft.margin.only(bottom=20)
                        ),
                        ft.Text(
                            "Nuevo Monto",
                            size=16,
                            weight=ft.FontWeight.BOLD,
                            color=ft.colors.BLUE_900
                        ),
                        self.edit_fee_field,
                    ],
                    spacing=15,
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER
                ),
                padding=20,
                bgcolor=ft.colors.WHITE,
                border_radius=12,
                shadow=ft.BoxShadow(
                    spread_radius=1,
                    blur_radius=10,
                    color=ft.colors.GREY_300,
                ),
                height=300  # Altura máxima del modal
            ),
            actions=[
                ft.TextButton(
                    "Cancelar",
                    on_click=self.close_edit_fee_modal,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.WHITE,
                        color=ft.colors.BLACK87,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
                ft.ElevatedButton(
                    "Guardar",
                    on_click=self.save_monthly_fee,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE_900,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True
        )
        self.page.overlay.append(self.edit_fee_modal)

        # Modal de éxito al guardar pago
        self.success_modal = ft.AlertDialog(
            title=ft.Text("¡Éxito!", size=20, weight=ft.FontWeight.BOLD),
            content=ft.Column(
                controls=[
                    ft.Icon(ft.icons.CHECK_CIRCLE, color=ft.colors.GREEN, size=40),
                    ft.Text("El pago se ha guardado correctamente", size=14),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=10,
                width=300,
                height=100,
            ),
            actions=[
                ft.ElevatedButton(
                    "Aceptar",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.GREEN_700,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=20, vertical=8),
                        text_style=ft.TextStyle(size=14, weight=ft.FontWeight.BOLD),
                    ),
                    on_click=lambda e: self.close_success_modal(e)
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.CENTER,
            modal=True,
        )
        self.page.overlay.append(self.success_modal)

        # Layout principal
        self.content = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Container(
                        content=self.title,
                        padding=ft.padding.only(bottom=30, top=0, left=10, right=10),
                        alignment=ft.alignment.top_left,
                        width=900,
                    ),
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                ft.Container(
                                    content=ft.Row(
                                        controls=[
                                            self.search_field,
                                            self.date_from_field,
                                            self.date_to_field,
                                            self.payment_method,
                                            self.status_filter,
                                            self.clear_btn,
                                            self.new_payment_btn,
                                        ],
                                        alignment=ft.MainAxisAlignment.START,
                                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                                        spacing=18,
                                        expand=True,
                                    ),
                                    padding=ft.padding.only(bottom=10, left=10),
                                    alignment=ft.alignment.top_left,
                                    expand=True,
                                ),
                            ],
                            alignment=ft.MainAxisAlignment.START,
                            spacing=20,
                        ),
                        padding=ft.padding.only(bottom=10, left=10),
                        alignment=ft.alignment.top_left,
                        width=1300,
                    ),
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                self.monthly_fee_card,  # Mover la tarjeta aquí
                                ft.Container(
                                    content=ft.Row(
                                        controls=[
                                            self.export_excel_btn,
                                            self.export_pdf_btn,
                                        ],
                                        alignment=ft.MainAxisAlignment.END,
                                        spacing=8,
                                    ),
                                    padding=ft.padding.only(bottom=10, left=10, right=30),
                                    alignment=ft.alignment.top_right,
                                ),
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            spacing=8,
                        ),
                        padding=ft.padding.only(bottom=10, left=10, right=30),
                        alignment=ft.alignment.top_right,
                        width=1300,
                    ),
                    ft.Container(
                        content=ft.Row(
                            controls=[
                                ft.Container(
                                    content=self.payments_table,
                                    expand=True,
                                    alignment=ft.alignment.center,
                                    padding=ft.padding.symmetric(horizontal=40),
                                    height=600,
                                )
                            ],
                            alignment=ft.MainAxisAlignment.CENTER,
                            vertical_alignment=ft.CrossAxisAlignment.START,
                            expand=True,
                        ),
                        alignment=ft.alignment.top_left,
                        width=1300,
                        padding=ft.padding.only(top=30),
                    ),
                    # Widget de paginación
                    ft.Container(
                        content=self.pagination_widget.get_widget(),
                        padding=ft.padding.only(top=20, bottom=20),
                        alignment=ft.alignment.center,
                    ),
                ],
                spacing=0,
                scroll=ft.ScrollMode.ALWAYS,
                horizontal_alignment=ft.CrossAxisAlignment.START,
            ),
            expand=True,
            padding=ft.padding.symmetric(horizontal=0, vertical=0),
            bgcolor=ft.colors.WHITE,
            border_radius=18,
            margin=ft.margin.symmetric(horizontal=0, vertical=0),
            alignment=ft.alignment.top_left,
        )

    def get_content(self):
        # Llamar load_data después de que la vista esté completamente inicializada
        self.page.loop.create_task(self._load_data_async())
        return self.content
    
    async def _load_data_async(self):
        """Carga los datos de forma asíncrona después de que la vista esté lista"""
        try:
            # Pequeño delay para asegurar que la vista esté completamente renderizada
            import asyncio
            await asyncio.sleep(0.1)
            
            print("[DEBUG - Pagos] Iniciando load_data asíncrono")
            
            # Cargar pagos usando session_scope para evitar problemas de sesión
            with session_scope() as session:
                # Recrear el controlador con la nueva sesión
                temp_controller = PaymentController(session)
                payments = [p for p in temp_controller.get_payments() if p.estado == 1]
                print(f"[DEBUG - Pagos] Obtenidos {len(payments)} pagos")
                
                # Actualizar paginación
                self.pagination_controller.set_items(payments)
                self.pagination_widget.update_items(payments)
                print("[DEBUG - Pagos] Paginación actualizada")
                
                # Actualizar tabla
                self.update_payments_table(payments)
                print("[DEBUG - Pagos] Tabla actualizada")
            
            # Cargar la cuota mensual actual
            try:
                with session_scope() as session:
                    current_fee = session.query(CuotaMensual).filter_by(activo=1).first()
                    self.current_monthly_fee = current_fee.monto if current_fee else None
            except Exception as e:
                print(f"[DEBUG - Pagos] Error al cargar cuota mensual: {str(e)}")
                self.current_monthly_fee = None

            # Cargar métodos de pago activos usando session_scope
            try:
                with session_scope() as session:
                    active_payment_methods = session.query(MetodoPago).filter_by(estado=True).all()
                    self.new_payment_method_field.options = [
                        ft.dropdown.Option(method.descripcion) for method in active_payment_methods
                    ]
            except Exception as e:
                print(f"[DEBUG - Pagos] Error al cargar métodos de pago: {str(e)}")
                
        except Exception as e:
            print(f"[DEBUG - Pagos] Error en load_data asíncrono: {str(e)}")
            self.update_payments_table([])
    
    def _on_page_change(self):
        """Callback cuando cambia la página"""
        self.update_payments_table()

    def create_summary_card(self, title: str, value: str, icon: str, color: str):
        return ft.Container(
            content=ft.Column(
                controls=[
                    ft.Icon(name=icon, color=color, size=32),
                    ft.Text(value, size=28, weight=ft.FontWeight.BOLD),
                    ft.Text(title, size=14, color=ft.colors.GREY_700),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=10,
            ),
            width=250,
            height=150,
            border_radius=10,
            bgcolor=ft.colors.WHITE,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=10,
                color=ft.colors.GREY_300,
            ),
            padding=20,
        )

    def show_new_payment_modal(self, e):
        if self.new_payment_modal not in self.page.overlay:
            self.page.overlay.append(self.new_payment_modal)
        self.new_payment_modal.open = True
        self.page.update()

    def close_modal(self, e):
        """
        Cierra el modal de nuevo pago y limpia los campos
        """
        self.new_payment_client_field.value = ""
        self.new_payment_date_value = None
        self.new_payment_date_picker.value = None
        self.new_payment_date_field.content.controls[0].value = "Seleccionar fecha"
        self.new_payment_amount_field.value = "0.00"
        self.new_payment_method_field.value = None
        self.new_payment_observations_field.value = ""
        self.selected_member_data = None
        self.member_search_results.visible = False
        self.member_search_results_container.height = 0
        self.amount_warning_text.visible = False
        self.receipt_validation_text.visible = False  # Ocultar el mensaje de validación
        self.new_payment_modal.open = False
        self.page.update()

    def load_data(self):
        """
        Carga los datos iniciales de la vista
        """
        # Cargar pagos usando session_scope para evitar problemas de sesión
        try:
            with session_scope() as session:
                # Recrear el controlador con la nueva sesión
                temp_controller = PaymentController(session)
                payments = [p for p in temp_controller.get_payments() if p.estado == 1]
                self.update_payments_table(payments)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al cargar datos: {str(e)}")
            self.update_payments_table([])

        # Cargar la cuota mensual actual
        try:
            with session_scope() as session:
                current_fee = session.query(CuotaMensual).filter_by(activo=1).first()
                self.current_monthly_fee = current_fee.monto if current_fee else None
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al cargar cuota mensual: {str(e)}")
            self.current_monthly_fee = None

        # Cargar métodos de pago activos usando session_scope
        try:
            with session_scope() as session:
                active_payment_methods = session.query(MetodoPago).filter_by(estado=True).all()
                self.new_payment_method_field.options = [
                    ft.dropdown.Option(method.descripcion) for method in active_payment_methods
                ]
                self.edit_payment_method_field.options = [
                    ft.dropdown.Option(method.descripcion) for method in active_payment_methods
                ]
                self.payment_method.options = [
                    ft.dropdown.Option("Todos")
                ] + [
                    ft.dropdown.Option(method.descripcion) for method in active_payment_methods
                ]
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al cargar métodos de pago: {str(e)}")
        
        self.page.update()

    def update_payments_table(self, payments=None):
        """
        Actualiza la tabla de pagos con datos reales
        """
        print(f"[DEBUG - Pagos] Actualizando tabla con {len(payments) if payments else 'None'} pagos")
        self.payments_table.rows.clear()
        
        # Obtener pagos de la página actual
        if payments is None:
            payments = self.pagination_controller.get_current_page_items()
            print(f"[DEBUG - Pagos] Pagos de página actual: {len(payments)}")
        
        if not payments:
            # Mostrar mensaje cuando no hay pagos
            self.payments_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Icon(name=ft.icons.PAYMENTS_OUTLINED, size=48, color=ft.colors.GREY_400),
                                        ft.Text(
                                            "No se encontraron pagos",
                                            size=20,
                                            weight=ft.FontWeight.BOLD,
                                            color=ft.colors.GREY_700
                                        ),
                                        ft.Text(
                                            "Registra tu primer pago usando el botón 'Nuevo Pago'",
                                            size=16,
                                            color=ft.colors.GREY_600
                                        ),
                                    ],
                                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                                    spacing=10,
                                ),
                                padding=40,
                                alignment=ft.alignment.center,
                            )
                        ),
                        ft.DataCell(ft.Container()),
                        ft.DataCell(ft.Container()),
                        ft.DataCell(ft.Container()),
                        ft.DataCell(ft.Container()),
                        ft.DataCell(ft.Container()),
                        ft.DataCell(ft.Container()),
                    ]
                )
            )
        else:
            for payment in payments:
                estado_texto = "Pagado" if payment.estado == 1 else "Cancelado"
                color_estado = ft.colors.GREEN if payment.estado == 1 else ft.colors.RED
                
                # Crear un objeto ligero con solo los datos necesarios para evitar problemas de sesión
                payment_data = type('PaymentData', (), {
                    'id_pago': payment.id_pago,
                    'fecha_pago': payment.fecha_pago,
                    'monto': payment.monto,
                    'referencia': payment.referencia,
                    'estado': payment.estado,
                    'miembro_nombre': f"{payment.miembro.nombre} {payment.miembro.apellido}",
                    'metodo_pago_descripcion': payment.metodo_pago.descripcion
                })
                
                self.payments_table.rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(payment_data.miembro_nombre)),
                            ft.DataCell(ft.Text(payment_data.fecha_pago.strftime("%d/%m/%Y"))),
                            ft.DataCell(ft.Text(f"${payment_data.monto}")),
                            ft.DataCell(ft.Text(payment_data.metodo_pago_descripcion)),
                            ft.DataCell(ft.Text(payment_data.referencia if payment_data.referencia else "")),
                            ft.DataCell(
                                ft.Container(
                                    content=ft.Text(
                                        estado_texto,
                                        color=color_estado
                                    ),
                                    bgcolor=ft.colors.GREY_100 if payment_data.estado == 1 else ft.colors.RED_100,
                                    border_radius=8,
                                    padding=5,
                                )
                            ),
                            ft.DataCell(
                                ft.Row(
                                    controls=[
                                        ft.IconButton(
                                            icon=ft.icons.EDIT,
                                            icon_color=ft.colors.BLUE,
                                            tooltip="Editar",
                                            on_click=lambda e, p=payment_data: self.edit_payment(p)
                                        ),
                                        ft.IconButton(
                                            icon=ft.icons.DELETE,
                                            icon_color=ft.colors.RED,
                                            tooltip="Eliminar",
                                            on_click=lambda e, p=payment_data: self.delete_payment(p)
                                        ),
                                    ],
                                    spacing=0,
                                )
                            ),
                        ]
                    )
                )
        self.page.update()

    def apply_filters(self, e=None):
        """
        Aplica los filtros seleccionados
        """
        filters = {}
        
        if self.search_field.value:
            filters['member_name'] = self.search_field.value
        
        # Filtro de fecha desde
        if self.date_from.value:
            filters['date_from'] = self.date_from.value
        
        # Filtro de fecha hasta
        if self.date_to.value:
            filters['date_to'] = self.date_to.value
        
        if self.payment_method.value and self.payment_method.value != "Todos":
            filters['payment_method'] = self.payment_method.value
        
        # Filtro de estado
        if self.status_filter.value == "Pagado":
            filters['estado'] = 1
        elif self.status_filter.value == "Cancelado":
            filters['estado'] = 0
        # Si es "Todos", no se agrega filtro
        
        try:
            with session_scope() as session:
                # Recrear el controlador con la nueva sesión
                temp_controller = PaymentController(session)
                payments = temp_controller.get_payments(filters)
                # Si el filtro es por estado, filtrar aquí también por si acaso
                if 'estado' in filters:
                    payments = [p for p in payments if p.estado == filters['estado']]
                
                # Actualizar paginación con los datos filtrados
                self.pagination_controller.set_items(payments)
                self.pagination_widget.update_items(payments)
                self.update_payments_table()
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al aplicar filtros: {str(e)}")
            self.update_payments_table([])

    def clear_filters(self, e):
        """
        Limpia todos los filtros
        """
        self.search_field.value = ""
        self.date_from.value = None
        self.date_to.value = None
        self.date_from_field.content.controls[0].value = ""
        self.date_to_field.content.controls[0].value = ""
        self.payment_method.value = "Todos"
        self.status_filter.value = "Pagado"
        self.page.update()
        # Recargar datos sin filtros usando el método asíncrono
        self.page.loop.create_task(self._load_data_async())

    def edit_payment(self, payment):
        """
        Abre el modal para editar un pago, cargando los datos del pago seleccionado
        """
        try:
            # Recargar el pago desde la base de datos con todas las relaciones
            with session_scope() as session:
                payment_fresh = session.query(Pago).options(
                    joinedload(Pago.miembro),
                    joinedload(Pago.metodo_pago)
                ).filter_by(id_pago=payment.id_pago).first()
                
                if not payment_fresh:
                    self.show_message("No se pudo encontrar el pago", ft.colors.RED)
                    return
                
                self.selected_payment = payment_fresh
                # Almacenar datos importantes para evitar problemas de sesión
                self.selected_payment_id = payment_fresh.id_pago
                self.selected_payment_member_id = payment_fresh.miembro.id_miembro
                
                self.edit_payment_client_field.value = f"{payment_fresh.miembro.nombre} {payment_fresh.miembro.apellido}"
                self.edit_payment_date_value = payment_fresh.fecha_pago
                self.edit_payment_date_picker.value = payment_fresh.fecha_pago
                self.edit_payment_date_field.content.controls[0].value = payment_fresh.fecha_pago.strftime("%d/%m/%Y")
                self.edit_payment_amount_field.value = str(payment_fresh.monto)
                self.edit_payment_method_field.value = payment_fresh.metodo_pago.descripcion
                self.edit_payment_observations_field.value = payment_fresh.referencia if payment_fresh.referencia else ""
                self.edit_payment_modal.open = True
                self.page.update()
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al cargar datos del pago: {str(e)}")
            self.show_message(f"Error al cargar los datos del pago: {str(e)}", ft.colors.RED)

    def close_edit_modal(self, e):
        self.edit_payment_client_field.value = ""
        self.edit_payment_date_value = None
        self.edit_payment_date_picker.value = None
        self.edit_payment_date_field.content.controls[0].value = "Seleccionar fecha"
        self.edit_payment_amount_field.value = "0.00"
        self.edit_payment_method_field.value = None
        self.edit_payment_observations_field.value = ""
        self.edit_payment_modal.open = False
        self.selected_payment = None
        # Limpiar también los IDs almacenados
        self.selected_payment_id = None
        self.selected_payment_member_id = None
        self.page.update()

    def on_edit_payment_date_change(self, e):
        self.edit_payment_date_value = self.edit_payment_date_picker.value
        value = self.edit_payment_date_picker.value.strftime("%d/%m/%Y") if self.edit_payment_date_picker.value else "Seleccionar fecha"
        self.edit_payment_date_field.content.controls[0].value = value
        self.page.update()

    def update_payment(self, e):
        if not hasattr(self, 'selected_payment_id') or not self.selected_payment_id:
            self.show_message("No hay pago seleccionado para editar", ft.colors.RED)
            return
        if not self.edit_payment_date_value:
            self.show_message("Debe seleccionar una fecha", ft.colors.RED)
            return
        if not self.edit_payment_amount_field.value or float(self.edit_payment_amount_field.value) <= 0:
            self.show_message("Debe ingresar un monto válido", ft.colors.RED)
            return
        if not self.edit_payment_method_field.value:
            self.show_message("Debe seleccionar un método de pago", ft.colors.RED)
            return
        
        payment_data = {
            'fecha_pago': self.edit_payment_date_value,
            'monto': float(self.edit_payment_amount_field.value),
            'id_miembro': self.selected_payment_member_id,  # Usar el ID almacenado
            'id_metodo_pago': self.get_payment_method_id(self.edit_payment_method_field.value),
            'referencia': self.edit_payment_observations_field.value
        }
        
        success, message = self.payment_controller.update_payment(self.selected_payment_id, payment_data)
        if success:
            self.show_message(message, ft.colors.GREEN)
            self.close_edit_modal(e)
            self.load_data()
        else:
            self.show_message(message, ft.colors.RED)

    def delete_payment(self, payment):
        """
        Abre el modal de confirmación para eliminar un pago
        """
        # Solo necesitamos el ID para eliminar, no las relaciones
        self.selected_payment_to_delete = payment
        self.delete_confirm_modal.open = True
        self.page.update()

    def close_delete_modal(self, e):
        self.delete_confirm_modal.open = False
        self.selected_payment_to_delete = None
        self.page.update()

    def confirm_delete_payment(self, e):
        if not self.selected_payment_to_delete:
            self.show_message("No hay pago seleccionado para cancelar", ft.colors.RED)
            self.delete_confirm_modal.open = False
            self.page.update()
            return
        # Cancelar pago: poner estado en 0
        payment_data = {'estado': 0}
        success, message = self.payment_controller.update_payment(self.selected_payment_to_delete.id_pago, payment_data)
        if success:
            self.show_message("Pago cancelado correctamente", ft.colors.GREEN)
            self.delete_confirm_modal.open = False
            self.selected_payment_to_delete = None
            self.load_data()
        else:
            self.show_message("Error al cancelar el pago", ft.colors.RED)
            self.delete_confirm_modal.open = False
            self.selected_payment_to_delete = None
        self.page.update()

    def generate_receipt(self, payment):
        """
        Prepara la generación del comprobante de pago
        """
        self.selected_payment_for_receipt = payment
        self.receipt_confirm_modal.open = True
        self.page.update()

    def close_receipt_modal(self, e):
        """
        Cierra el modal de confirmación del comprobante
        """
        self.receipt_confirm_modal.open = False
        self.selected_payment_for_receipt = None
        self.page.update()

    def confirm_generate_receipt(self, e):
        """
        Confirma la generación del comprobante
        """
        if not self.selected_payment_for_receipt:
            self.receipt_validation_text.value = "No hay pago seleccionado"
            self.receipt_validation_text.color = ft.colors.RED
            self.receipt_validation_text.visible = True
            self.page.update()
            return

        try:
            import logging
            logger = logging.getLogger(__name__)
            logger.info("Iniciando generación de comprobante...")
            
            # Crear el documento PDF
            downloads_path = os.path.expanduser("~/Downloads")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_path = os.path.join(downloads_path, f"comprobante_pago_{timestamp}.pdf")
            
            doc = SimpleDocTemplate(
                pdf_path,
                pagesize=letter,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )

            # Crear los estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#1F4E78'),
                alignment=1  # Centrado
            )

            # Crear el contenido
            elements = []

            # Título
            elements.append(Paragraph("Comprobante de Pago", title_style))
            elements.append(Spacer(1, 20))

            # Información del pago
            payment_info = [
                ["Miembro:", f"{self.selected_payment_for_receipt.miembro.nombre} {self.selected_payment_for_receipt.miembro.apellido}"],
                ["Fecha:", self.selected_payment_for_receipt.fecha_pago.strftime("%d/%m/%Y")],
                ["Monto:", f"${self.selected_payment_for_receipt.monto:,.2f}"],
                ["Método de Pago:", self.selected_payment_for_receipt.metodo_pago.descripcion],
                ["Referencia:", self.selected_payment_for_receipt.referencia or "-"]
            ]

            # Crear la tabla de información
            table = Table(payment_info, colWidths=[150, 350])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 30))

            # Pie de página
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=1
            )
            footer = Paragraph(
                f"Comprobante generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                footer_style
            )
            elements.append(footer)

            # Construir el PDF
            doc.build(elements)

            # Leer el contenido del PDF generado
            with open(pdf_path, 'rb') as pdf_file:
                pdf_content = pdf_file.read()

            # Si es un pago temporal (desde el formulario), no guardamos en la base de datos
            if hasattr(self.selected_payment_for_receipt, 'id_pago') and self.selected_payment_for_receipt.id_pago > 0:
                # Guardar el comprobante en la base de datos
                success, message = self.payment_controller.save_payment_receipt(
                    self.selected_payment_for_receipt.id_pago,
                    pdf_content
                )
                if not success:
                    logger.error(f"Error al guardar comprobante: {message}")
                    self.receipt_validation_text.value = f"Error al guardar el comprobante: {message}"
                    self.receipt_validation_text.color = ft.colors.RED
                else:
                    self.receipt_validation_text.value = "Comprobante generado y guardado exitosamente"
                    self.receipt_validation_text.color = ft.colors.GREEN
            else:
                self.receipt_validation_text.value = "Comprobante generado exitosamente, se ha guardado en la carpeta de descargas."
                self.receipt_validation_text.color = ft.colors.GREEN
            
            self.receipt_validation_text.visible = True
            
            # Cerrar el modal
            self.close_receipt_modal(e)

        except Exception as e:
            logger.error(f"Error al generar comprobante: {str(e)}")
            
            # Mostrar mensaje de error
            self.receipt_validation_text.value = f"Error al generar el comprobante: {str(e)}"
            self.receipt_validation_text.color = ft.colors.RED
            self.receipt_validation_text.visible = True
            self.page.update()

    def search_member(self, e):
        """
        Busca miembros según el texto ingresado
        """
        search_text = self.new_payment_client_field.value.lower()
        if not search_text:
            self.member_search_results.visible = False
            self.member_search_results_container.height = 0
            self.page.update()
            return

        # Buscar miembros que coincidan con el texto de búsqueda usando session_scope
        try:
            with session_scope() as session:
                members_query = session.query(Miembro).filter(
                    (Miembro.nombre.ilike(f"%{search_text}%")) |
                    (Miembro.apellido.ilike(f"%{search_text}%")) |
                    (Miembro.documento.ilike(f"%{search_text}%"))
                ).limit(5).all()
                
                # Extraer todos los datos inmediatamente mientras la sesión está activa
                members_data = []
                for member in members_query:
                    member_data = type('MemberData', (), {
                        'id_miembro': member.id_miembro,
                        'nombre': member.nombre,
                        'apellido': member.apellido,
                        'documento': member.documento
                    })
                    members_data.append(member_data)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al buscar miembros: {str(e)}")
            members_data = []

        self.member_search_results.controls.clear()
        
        if members_data:
            for member_data in members_data:
                self.member_search_results.controls.append(
                    ft.Container(
                        content=ft.ListTile(
                            leading=ft.Icon(ft.icons.PERSON),
                            title=ft.Text(f"{member_data.nombre} {member_data.apellido}"),
                            subtitle=ft.Text(f"Documento: {member_data.documento}"),
                            on_click=lambda e, m=member_data: self.select_member(m)
                        ),
                        border=ft.border.all(1, ft.colors.GREY_300),
                        border_radius=8,
                        bgcolor=ft.colors.WHITE,
                    )
                )
            self.member_search_results.visible = True
            self.member_search_results_container.height = 60
        else:
            self.member_search_results.controls.append(
                ft.Container(
                    content=ft.Text("No se encontraron miembros", color=ft.colors.GREY_700),
                    padding=10,
                    alignment=ft.alignment.center,
                )
            )
            self.member_search_results.visible = True
            self.member_search_results_container.height = 60
        
        self.page.update()

    def select_member(self, member):
        """
        Selecciona un miembro de la lista de resultados
        """
        # Guardar solo los datos necesarios del miembro
        self.selected_member_data = {
            'id': member.id_miembro,
            'nombre': member.nombre,
            'apellido': member.apellido,
            'documento': member.documento
        }
        self.new_payment_client_field.value = f"{member.nombre} {member.apellido}"
        self.member_search_results.visible = False
        self.member_search_results_container.height = 0
        self.page.update()

    def save_payment(self, e):
        """
        Guarda un nuevo pago
        """
        if not self.selected_member_data:
            self.show_message("Debe seleccionar un miembro", ft.colors.RED)
            return

        if not self.new_payment_date_value:
            self.show_message("Debe seleccionar una fecha", ft.colors.RED)
            return

        try:
            monto = float(self.new_payment_amount_field.value)
            if monto <= 0:
                self.show_message("Debe ingresar un monto válido", ft.colors.RED)
                return
        except ValueError:
            self.show_message("El monto ingresado no es válido", ft.colors.RED)
            return

        if not self.new_payment_method_field.value:
            self.show_message("Debe seleccionar un método de pago", ft.colors.RED)
            return

        try:
            # Limpiar cualquier transacción pendiente
            if hasattr(self, 'db_session'):
                try:
                    self.db_session.rollback()
                except:
                    pass
                self.db_session.close()
            
            logger.info("Obteniendo ID del método de pago...")
            metodo_pago_id = self.get_payment_method_id(self.new_payment_method_field.value)
            if not metodo_pago_id:
                self.show_message("Error al obtener el método de pago", ft.colors.RED)
                return

            logger.info("Preparando datos del pago...")
            payment_data = {
                'fecha_pago': self.new_payment_date_value,
                'monto': monto,
                'id_miembro': self.selected_member_data['id'],
                'id_metodo_pago': metodo_pago_id,
                'referencia': self.new_payment_observations_field.value
            }
            
            logger.info("Llamando al controlador para crear el pago...")
            success, response = self.payment_controller.create_payment(payment_data)
            
            if success:
                logger.info("Pago creado exitosamente")
                payment_id = response['id_pago']  # Obtener el ID del pago creado
                
                # Generar el comprobante para el pago recién creado
                try:
                    # Crear el documento PDF
                    downloads_path = os.path.expanduser("~/Downloads")
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    pdf_path = os.path.join(downloads_path, f"comprobante_pago_{timestamp}.pdf")
                    
                    doc = SimpleDocTemplate(
                        pdf_path,
                        pagesize=letter,
                        rightMargin=30,
                        leftMargin=30,
                        topMargin=30,
                        bottomMargin=30
                    )

                    # Crear los estilos
                    styles = getSampleStyleSheet()
                    title_style = ParagraphStyle(
                        'CustomTitle',
                        parent=styles['Heading1'],
                        fontSize=24,
                        spaceAfter=30,
                        textColor=colors.HexColor('#1F4E78'),
                        alignment=1  # Centrado
                    )

                    # Crear el contenido
                    elements = []

                    # Título
                    elements.append(Paragraph("Comprobante de Pago", title_style))
                    elements.append(Spacer(1, 20))

                    # Información del pago
                    payment_info = [
                        ["Miembro:", f"{self.selected_member_data['nombre']} {self.selected_member_data['apellido']}"],
                        ["Fecha:", self.new_payment_date_value.strftime("%d/%m/%Y")],
                        ["Monto:", f"${monto:,.2f}"],
                        ["Método de Pago:", self.new_payment_method_field.value],
                        ["Referencia:", self.new_payment_observations_field.value or "-"]
                    ]

                    # Crear la tabla de información
                    table = Table(payment_info, colWidths=[150, 350])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1F4E78')),
                        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, -1), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                        ('TOPPADDING', (0, 0), (-1, -1), 12),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
                    ]))

                    elements.append(table)
                    elements.append(Spacer(1, 30))

                    # Pie de página
                    footer_style = ParagraphStyle(
                        'Footer',
                        parent=styles['Normal'],
                        fontSize=8,
                        textColor=colors.gray,
                        alignment=1
                    )
                    footer = Paragraph(
                        f"Comprobante generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                        footer_style
                    )
                    elements.append(footer)

                    # Construir el PDF
                    doc.build(elements)

                    # Leer el contenido del PDF generado
                    with open(pdf_path, 'rb') as pdf_file:
                        pdf_content = pdf_file.read()

                    # Guardar el comprobante en la base de datos
                    success, message = self.payment_controller.save_payment_receipt(
                        payment_id,  # Usar el ID del pago recién creado
                        pdf_content
                    )
                    if not success:
                        logger.error(f"Error al guardar comprobante: {message}")
                        self.show_message(f"Error al guardar el comprobante: {message}", ft.colors.RED)
                    else:
                        logger.info("Comprobante guardado exitosamente en la base de datos")

                except Exception as ex:
                    logger.error(f"Error al generar comprobante automático: {str(ex)}")
                    self.show_message(f"Error al generar el comprobante: {str(ex)}", ft.colors.RED)
                
                self.close_modal(e)
                self.load_data()
                # Mostrar modal de éxito
                self.success_modal.open = True
                self.page.update()
            else:
                logger.error(f"Error al crear el pago: {response}")
                self.show_message(response, ft.colors.RED)
        except Exception as e:
            logger.error(f"Error inesperado al guardar el pago: {str(e)}")
            self.show_message(f"Error al guardar el pago: {str(e)}", ft.colors.RED)

    def get_payment_method_id(self, method_name):
        """
        Obtiene el ID del método de pago según su nombre
        """
        try:
            with session_scope() as session:
                method = session.query(MetodoPago).filter_by(descripcion=method_name).first()
                return method.id_metodo_pago if method else None
        except Exception as e:
            logger.error(f"Error al obtener método de pago: {str(e)}")
            return None

    def show_message(self, message: str, color: str):
        self.page.snack_bar = ft.SnackBar(
            content=ft.Text(message),
            bgcolor=color,
            duration=5000 if color == ft.colors.ORANGE else 3000,  # 5 segundos para advertencias, 3 para otros mensajes
            action="OK",
            action_color=ft.colors.WHITE
        )
        self.page.snack_bar.open = True
        self.page.update()

    def on_date_from_change(self, e):
        self.date_from_value = self.date_from.value
        value = self.date_from.value.strftime("%d/%m/%Y") if self.date_from.value else ""
        self.date_from_field.content.controls[0].value = value
        self.page.update()

    def on_date_to_change(self, e):
        self.date_to_value = self.date_to.value
        value = self.date_to.value.strftime("%d/%m/%Y") if self.date_to.value else ""
        self.date_to_field.content.controls[0].value = value
        self.page.update()
        self.apply_filters()  # Aplicar filtros automáticamente al cambiar la fecha hasta

    def open_date_picker(self, picker):
        picker.open = True
        self.page.update()

    def on_new_payment_date_change(self, e):
        self.new_payment_date_value = self.new_payment_date_picker.value
        value = self.new_payment_date_picker.value.strftime("%d/%m/%Y") if self.new_payment_date_picker.value else "Seleccionar fecha"
        self.new_payment_date_field.content.controls[0].value = value
        self.page.update()

    def show_success_dialog(self, file_path: str):
        """
        Muestra un diálogo de éxito con un botón para abrir la carpeta
        """
        def open_folder(e):
            folder_path = os.path.dirname(file_path)
            if os.name == 'nt':  # Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # macOS y Linux
                subprocess.run(['xdg-open', folder_path])
            success_dialog.open = False
            self.page.update()

        success_dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.icons.CHECK_CIRCLE, color=ft.colors.GREEN, size=30),
                ft.Text("¡Descargado con éxito!", size=20, weight=ft.FontWeight.BOLD),
            ]),
            content=ft.Column([
                ft.Text(f"El archivo se ha guardado en:", size=14),
                ft.Text(file_path, size=14, color=ft.colors.GREY_700),
                ft.Container(height=20),
                ft.ElevatedButton(
                    "Abrir carpeta",
                    icon=ft.icons.FOLDER_OPEN,
                    on_click=open_folder,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE,
                        color=ft.colors.WHITE,
                    ),
                ),
            ], tight=True, spacing=10),
            actions=[
                ft.TextButton("Cerrar", on_click=lambda e: self.close_success_dialog(success_dialog)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        if success_dialog not in self.page.overlay:
            self.page.overlay.append(success_dialog)
        success_dialog.open = True
        self.page.update()

    def close_success_dialog(self, dialog):
        """
        Cierra el diálogo de éxito
        """
        dialog.open = False
        self.page.update()

    def export_to_excel(self, e):
        """
        Exporta los datos a Excel
        """
        try:
            with session_scope() as session:
                temp_controller = PaymentController(session)
                payments = temp_controller.get_payments()
                file_path = self._export_to_excel(payments, os.path.expanduser("~/Downloads"))
                self.show_success_dialog(file_path)
        except Exception as ex:
            self.show_message(
                ft.Row([
                    ft.Icon(ft.icons.ERROR, color=ft.colors.RED),
                    ft.Text(f"Error al exportar a Excel: {str(ex)}", color=ft.colors.RED)
                ]),
                ft.colors.RED_50
            )

    def export_to_pdf(self, e):
        """
        Exporta los datos a PDF
        """
        try:
            with session_scope() as session:
                temp_controller = PaymentController(session)
                payments = temp_controller.get_payments()
                file_path = self._export_to_pdf(payments, os.path.expanduser("~/Downloads"))
                self.show_success_dialog(file_path)
        except Exception as ex:
            self.show_message(
                ft.Row([
                    ft.Icon(ft.icons.ERROR, color=ft.colors.RED),
                    ft.Text(f"Error al exportar a PDF: {str(ex)}", color=ft.colors.RED)
                ]),
                ft.colors.RED_50
            )

    def check_overdue_payments(self):
        """
        Verifica los pagos vencidos y muestra una alerta
        """
        try:
            # Obtener todos los miembros usando session_scope
            with session_scope() as session:
                # Cargar miembros con sus pagos y métodos de pago eagerly
                members = session.query(Miembro).all()
                overdue_members = []

                for member in members:
                    # Obtener el último pago del miembro con carga eager
                    last_payment = session.query(Pago).options(
                        joinedload(Pago.metodo_pago)
                    ).filter(
                        Pago.id_miembro == member.id_miembro,
                        Pago.estado == 1  # Solo pagos activos
                    ).order_by(Pago.fecha_pago.desc()).first()

                    if last_payment:
                        # Calcular días desde el último pago
                        days_since_payment = (datetime.now() - last_payment.fecha_pago).days
                        
                        # Si han pasado más de 30 días, agregar a la lista de vencidos
                        if days_since_payment > 30:
                            overdue_members.append({
                                'member_name': f"{member.nombre} {member.apellido}",
                                'days_overdue': days_since_payment,
                                'last_payment_date': last_payment.fecha_pago,
                                'last_payment_amount': last_payment.monto,
                                'payment_method': last_payment.metodo_pago.descripcion
                            })

            if overdue_members:
                # Crear el contenido de la alerta
                alert_content = ft.Column(
                    controls=[
                        # Encabezado con icono
                        ft.Container(
                            content=ft.Row(
                                controls=[
                                    ft.Icon(
                                        name=ft.icons.WARNING_ROUNDED,
                                        size=40,
                                        color=ft.colors.RED_700
                                    ),
                                    ft.Column(
                                        controls=[
                                            ft.Text(
                                                "¡Atención! Pagos Vencidos",
                                                size=24,
                                                weight=ft.FontWeight.BOLD,
                                                color=ft.colors.RED_700
                                            ),
                                            ft.Text(
                                                "Los siguientes miembros tienen pagos pendientes:",
                                                size=16,
                                                color=ft.colors.GREY_700
                                            )
                                        ],
                                        spacing=5
                                    )
                                ],
                                alignment=ft.MainAxisAlignment.START,
                                spacing=15
                            ),
                            padding=ft.padding.only(bottom=20)
                        ),
                        
                        # Lista de miembros con pagos vencidos
                        ft.Container(
                            content=ft.Column(
                                controls=[
                                    ft.Container(
                                        content=ft.Column(
                                            controls=[
                                                # Encabezado de la tarjeta
                                                ft.Row(
                                                    controls=[
                                                        ft.Icon(
                                                            name=ft.icons.PERSON,
                                                            color=ft.colors.BLUE_900,
                                                            size=24
                                                        ),
                                                        ft.Text(
                                                            m['member_name'],
                                                            size=16,
                                                            weight=ft.FontWeight.BOLD,
                                                            color=ft.colors.BLUE_900
                                                        ),
                                                        ft.Container(
                                                            content=ft.Text(
                                                                f"Vencido hace {m['days_overdue']} días",
                                                                color=ft.colors.WHITE,
                                                                weight=ft.FontWeight.BOLD,
                                                                size=12
                                                            ),
                                                            bgcolor=ft.colors.RED_700,
                                                            padding=ft.padding.symmetric(horizontal=12, vertical=6),
                                                            border_radius=20
                                                        )
                                                    ],
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                                                ),
                                                
                                                # Línea separadora
                                                ft.Container(
                                                    content=ft.Divider(height=1, color=ft.colors.GREY_300),
                                                    margin=ft.margin.symmetric(vertical=10)
                                                ),
                                                
                                                # Detalles del pago
                                                ft.Row(
                                                    controls=[
                                                        ft.Column(
                                                            controls=[
                                                                ft.Text(
                                                                    "Último Pago",
                                                                    size=12,
                                                                    color=ft.colors.GREY_600
                                                                ),
                                                                ft.Text(
                                                                    m['last_payment_date'].strftime("%d/%m/%Y"),
                                                                    size=14,
                                                                    weight=ft.FontWeight.BOLD
                                                                )
                                                            ],
                                                            spacing=2
                                                        ),
                                                        ft.Column(
                                                            controls=[
                                                                ft.Text(
                                                                    "Monto",
                                                                    size=12,
                                                                    color=ft.colors.GREY_600
                                                                ),
                                                                ft.Text(
                                                                    f"${m['last_payment_amount']:,.2f}",
                                                                    size=14,
                                                                    weight=ft.FontWeight.BOLD
                                                                )
                                                            ],
                                                            spacing=2
                                                        ),
                                                        ft.Column(
                                                            controls=[
                                                                ft.Text(
                                                                    "Método",
                                                                    size=12,
                                                                    color=ft.colors.GREY_600
                                                                ),
                                                                ft.Text(
                                                                    m['payment_method'],
                                                                    size=14,
                                                                    weight=ft.FontWeight.BOLD
                                                                )
                                                            ],
                                                            spacing=2
                                                        )
                                                    ],
                                                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                                                )
                                            ],
                                            spacing=5
                                        ),
                                        border=ft.border.all(1, ft.colors.GREY_300),
                                        border_radius=12,
                                        padding=15,
                                        bgcolor=ft.colors.WHITE,
                                        shadow=ft.BoxShadow(
                                            spread_radius=1,
                                            blur_radius=10,
                                            color=ft.colors.GREY_300,
                                        ),
                                        margin=ft.margin.only(bottom=10)
                                    ) for m in overdue_members
                                ],
                                scroll=ft.ScrollMode.AUTO,
                                height=300
                            ),
                            margin=ft.margin.only(top=10, bottom=10)
                        ),
                        
                        # Pie de página
                        ft.Container(
                            content=ft.Column(
                                controls=[
                                    ft.Text(
                                        "Acción Requerida",
                                        size=16,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.colors.BLUE_900
                                    ),
                                    ft.Text(
                                        "Por favor, contacte a estos miembros para regularizar sus pagos lo antes posible.",
                                        size=14,
                                        color=ft.colors.GREY_700
                                    )
                                ],
                                spacing=5,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                            padding=ft.padding.symmetric(vertical=10),
                            bgcolor=ft.colors.BLUE_50,
                            border_radius=8
                        )
                    ],
                    spacing=10,
                    width=600
                )

                # Crear el diálogo de alerta
                self.overdue_alert = ft.AlertDialog(
                    title=None,  # Quitamos el título por defecto
                    content=alert_content,
                    actions=[
                        ft.ElevatedButton(
                            "Entendido",
                            icon=ft.icons.CHECK_CIRCLE,
                            on_click=self.close_overdue_alert,
                            style=ft.ButtonStyle(
                                bgcolor=ft.colors.BLUE_900,
                                color=ft.colors.WHITE,
                                shape=ft.RoundedRectangleBorder(radius=8),
                                padding=ft.padding.symmetric(horizontal=28, vertical=12),
                                text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                            )
                        ),
                    ],
                    actions_alignment=ft.MainAxisAlignment.END,
                    modal=True
                )

                # Mostrar la alerta
                self.page.overlay.append(self.overdue_alert)
                self.overdue_alert.open = True
                self.page.update()

        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error al verificar pagos vencidos: {str(e)}")
            # No mostrar el error al usuario para no interrumpir la experiencia

    def close_overdue_alert(self, e):
        """
        Cierra la alerta de pagos vencidos
        """
        if hasattr(self, 'overdue_alert'):
            self.overdue_alert.open = False
            self.page.update()

    def show_edit_fee_modal(self, e):
        """
        Muestra el modal para editar la cuota mensual
        """
        current_fee = self.monthly_fee_controller.get_current_fee()
        if current_fee:
            self.edit_fee_field.value = str(current_fee.monto)
        self.edit_fee_modal.open = True
        self.page.update()

    def close_edit_fee_modal(self, e):
        """
        Cierra el modal de edición de cuota mensual
        """
        self.edit_fee_modal.open = False
        self.page.update()

    def save_monthly_fee(self, e):
        """
        Guarda la nueva cuota mensual
        """
        try:
            new_amount = float(self.edit_fee_field.value)
            if new_amount <= 0:
                self.show_message("El monto debe ser mayor a 0", ft.colors.RED)
                return

            success, message = self.monthly_fee_controller.update_fee(new_amount)
            if success:
                self.show_message(message, ft.colors.GREEN)
                # Actualizar el valor mostrado en la tarjeta
                self.monthly_fee_card.content.controls[0].controls[1].controls[1].value = f"${new_amount:,.2f}"
                # Actualizar la cuota mensual en memoria
                self.current_monthly_fee = new_amount
                self.close_edit_fee_modal(e)
                self.page.update()
            else:
                self.show_message(message, ft.colors.RED)
        except ValueError:
            self.show_message("Por favor, ingrese un monto válido", ft.colors.RED)

    def on_edit_fee_date_change(self, e):
        """
        Maneja el cambio de fecha en el selector de fecha de la cuota mensual
        """
        self.edit_fee_date_value = self.edit_fee_date_picker.value
        value = self.edit_fee_date_picker.value.strftime("%d/%m/%Y") if self.edit_fee_date_picker.value else "Seleccionar fecha"
        self.edit_fee_date_field.content.controls[0].value = value
        self.page.update()

    def validate_payment_amount(self, e):
        """
        Valida el monto ingresado contra la cuota mensual
        """
        try:
            # Solo validar si el valor es un número válido
            amount = float(self.new_payment_amount_field.value)
            
            # Usar la cuota mensual almacenada en memoria
            if self.current_monthly_fee and amount != self.current_monthly_fee:
                self.amount_warning_text.value = f"Advertencia: El monto ingresado (${amount:,.2f}) es diferente a la cuota mensual (${self.current_monthly_fee:,.2f})"
                self.amount_warning_text.visible = True
            else:
                self.amount_warning_text.visible = False
                
            self.page.update()
        except ValueError:
            # Si el valor no es un número válido, ocultamos la advertencia
            self.amount_warning_text.visible = False
            self.page.update()

    def generate_receipt_from_form(self, e):
        """
        Genera un comprobante con los datos ingresados en el formulario
        """
        if not self.selected_member_data:
            self.show_message("Debe seleccionar un miembro", ft.colors.RED)
            return

        if not self.new_payment_date_value:
            self.show_message("Debe seleccionar una fecha", ft.colors.RED)
            return

        try:
            monto = float(self.new_payment_amount_field.value)
            if monto <= 0:
                self.show_message("Debe ingresar un monto válido", ft.colors.RED)
                return
        except ValueError:
            self.show_message("El monto ingresado no es válido", ft.colors.RED)
            return

        if not self.new_payment_method_field.value:
            self.show_message("Debe seleccionar un método de pago", ft.colors.RED)
            return

        try:
            # Crear un objeto temporal con los datos del formulario
            temp_payment = type('TempPayment', (), {
                'id_pago': 0,  # ID temporal
                'fecha_pago': self.new_payment_date_value,
                'monto': monto,
                'miembro': type('TempMember', (), {
                    'nombre': self.selected_member_data['nombre'],
                    'apellido': self.selected_member_data['apellido'],
                    'documento': self.selected_member_data.get('documento', '')
                }),
                'metodo_pago': type('TempPaymentMethod', (), {
                    'descripcion': self.new_payment_method_field.value
                }),
                'referencia': self.new_payment_observations_field.value,
                'estado': 1
            })

            # Guardar el pago temporal para la generación del comprobante
            self.selected_payment_for_receipt = temp_payment
            
            # Generar el comprobante
            self.confirm_generate_receipt(e)

        except Exception as e:
            self.show_message(f"Error al generar el comprobante: {str(e)}", ft.colors.RED)

    def close_success_modal(self, e):
        """
        Cierra el modal de éxito
        """
        self.success_modal.open = False
        self.page.update()

    def _export_to_excel(self, payments, downloads_path):
        """
        Realiza la exportación a Excel
        """
        try:
            # Crear un nuevo libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pagos"

            # Estilos para los encabezados
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000")
            )

            # Estilos para las celdas de datos
            data_font = Font(size=11)
            data_alignment = Alignment(horizontal="center", vertical="center")
            data_border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000")
            )
            alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            # Agregar encabezados
            headers = ["Miembro", "Fecha", "Monto", "Método de Pago", "Estado"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border

            # Agregar datos
            for row, payment in enumerate(payments, 2):
                # Aplicar estilo alternado a las filas
                if row % 2 == 0:
                    row_fill = alternate_fill
                else:
                    row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                # Miembro
                cell = ws.cell(row=row, column=1, value=f"{payment.miembro.nombre} {payment.miembro.apellido}")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                cell.fill = row_fill

                # Fecha
                cell = ws.cell(row=row, column=2, value=payment.fecha_pago.strftime("%d/%m/%Y"))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                cell.fill = row_fill

                # Monto
                cell = ws.cell(row=row, column=3, value=payment.monto)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                cell.fill = row_fill
                cell.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

                # Método de Pago
                cell = ws.cell(row=row, column=4, value=payment.metodo_pago.descripcion)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                cell.fill = row_fill

                # Estado
                cell = ws.cell(row=row, column=5, value="Activo" if payment.estado == 1 else "Inactivo")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                cell.fill = row_fill

            # Ajustar el ancho de las columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Asegurar que la columna de monto tenga suficiente ancho
            ws.column_dimensions['C'].width = 15

            # Guardar el archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(downloads_path, f"pagos_{timestamp}.xlsx")
            wb.save(file_path)
            return file_path
        except Exception as ex:
            raise Exception(f"Error al exportar a Excel: {str(ex)}")

    def _export_to_pdf(self, payments, downloads_path):
        """
        Realiza la exportación a PDF
        """
        try:
            # Crear el documento PDF
            doc = SimpleDocTemplate(
                os.path.join(downloads_path, f"pagos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"),
                pagesize=letter,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )

            # Crear los estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#1F4E78'),
                alignment=1  # Centrado
            )
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=12
            )

            # Crear el contenido
            elements = []

            # Título
            elements.append(Paragraph("Reporte de Pagos", title_style))
            elements.append(Spacer(1, 20))

            # Tabla de datos
            data = [["Miembro", "Fecha", "Monto", "Método de Pago", "Estado"]]
            for payment in payments:
                data.append([
                    f"{payment.miembro.nombre} {payment.miembro.apellido}",
                    payment.fecha_pago.strftime("%d/%m/%Y"),
                    f"${payment.monto:,.2f}",
                    payment.metodo_pago.descripcion,
                    "Activo" if payment.estado == 1 else "Inactivo"
                ])

            # Crear la tabla
            table = Table(data)
            table.setStyle(TableStyle([
                # Estilo del encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Estilo de las filas de datos
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                
                # Bordes y líneas
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1F4E78')),
                
                # Alineación
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(table)

            # Pie de página con fecha y hora
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=1
            )
            footer = Paragraph(
                f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                footer_style
            )
            elements.append(Spacer(1, 20))
            elements.append(footer)

            # Construir el PDF
            doc.build(elements)
            return doc.filename
        except Exception as ex:
            raise Exception(f"Error al exportar a PDF: {str(ex)}")
