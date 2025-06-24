import flet as ft
from datetime import datetime
from gym_manager.controllers.member_controller import MemberController
from gym_manager.controllers.routine_controller import RoutineController
from gym_manager.utils.database import get_db_session
from gym_manager.services.excel_utils import export_members_to_excel, export_members_to_pdf
import os
import subprocess
import tempfile
from gym_manager.views.module_views import ModuleView
from gym_manager.models.member import Miembro
from pathlib import Path

class MembersView(ModuleView):
    def __init__(self, page: ft.Page):
        super().__init__(page, "Gestión de Miembros")
        self.member_controller = MemberController(get_db_session())
        self.routine_controller = RoutineController()
        self.db_session = get_db_session()
        
        # Inicializar referencias
        self.new_member_name = ft.Ref[ft.TextField]()
        self.new_member_lastname = ft.Ref[ft.TextField]()
        self.new_member_document = ft.Ref[ft.TextField]()
        self.new_member_email = ft.Ref[ft.TextField]()
        self.new_member_birth_date = ft.Ref[ft.ElevatedButton]()
        self.new_member_gender = ft.Ref[ft.Dropdown]()
        self.new_member_phone = ft.Ref[ft.TextField]()
        self.new_member_address = ft.Ref[ft.TextField]()
        self.new_member_membership = ft.Ref[ft.Dropdown]()
        self.new_member_start_date = ft.Ref[ft.ElevatedButton]()
        self.new_member_medical = ft.Ref[ft.TextField]()
        self.new_member_status = ft.Ref[ft.Dropdown]()
        
        # DatePickers
        self.birth_date_picker = ft.DatePicker(
            first_date=datetime(1950, 1, 1),
            last_date=datetime.now(),
            on_change=self.on_birth_date_change
        )
        self.start_date_picker = ft.DatePicker(
            first_date=datetime.now(),
            last_date=datetime(2100, 12, 31),
            on_change=self.on_start_date_change
        )
        
        # Agregar DatePickers al overlay de la página
        self.page.overlay.append(self.birth_date_picker)
        self.page.overlay.append(self.start_date_picker)
        
        self.setup_member_view()
        self.load_data()

    def setup_member_view(self):
        # Título amigable
        self.welcome_title = ft.Text(
            "¡Administra y consulta tus miembros!",
            size=28,
            weight=ft.FontWeight.BOLD,
            color=ft.colors.BLACK,
            text_align=ft.TextAlign.LEFT,
        )

        # Botón Nuevo Miembro
        self.new_member_btn = ft.ElevatedButton(
            text="Nuevo Miembro",
            icon=ft.icons.PERSON_ADD,
            style=ft.ButtonStyle(
                bgcolor=ft.colors.BLUE,
                color=ft.colors.WHITE,
                shape=ft.RoundedRectangleBorder(radius=10),
                padding=ft.padding.symmetric(horizontal=10, vertical=8),
                text_style=ft.TextStyle(size=15, weight=ft.FontWeight.BOLD),
            ),
            width=150,
            on_click=self.show_new_member_modal
        )

        # Botones de exportación
        self.export_excel_btn = ft.IconButton(
            icon=ft.icons.TABLE_VIEW,
            icon_color=ft.colors.GREEN_700,
            tooltip="Exportar a Excel",
            on_click=self.export_to_excel,
            width=48,
            height=48,
        )
        self.export_pdf_btn = ft.IconButton(
            icon=ft.icons.PICTURE_AS_PDF,
            icon_color=ft.colors.RED_700,
            tooltip="Exportar a PDF",
            on_click=self.export_to_pdf,
            width=48,
            height=48,
        )

        # Filtros de búsqueda
        self.search_field = ft.TextField(
            label="Buscar por nombre, documento o email",
            prefix_icon=ft.icons.SEARCH,
            border_radius=10,
            width=320,
            height=48,
            text_size=16,
            on_change=self.apply_filters
        )

        self.status_filter = ft.Dropdown(
            label="Estado",
            width=150,
            options=[
                ft.dropdown.Option("Todos"),
                ft.dropdown.Option("Activo"),
                ft.dropdown.Option("Inactivo")
            ],
            border_radius=10,
            text_size=16,
            on_change=self.apply_filters
        )

        self.membership_type = ft.Dropdown(
            label="Tipo de membresía",
            width=200,
            options=[
                ft.dropdown.Option("Todos"),
                ft.dropdown.Option("Mensual"),
                ft.dropdown.Option("Trimestral"),
                ft.dropdown.Option("Semestral"),
                ft.dropdown.Option("Anual")
            ],
            border_radius=10,
            text_size=16,
            on_change=self.apply_filters
        )

        self.clear_btn = ft.OutlinedButton(
            text="Limpiar filtros",
            icon=ft.icons.CLEAR,
            style=ft.ButtonStyle(
                color=ft.colors.GREY_700,
                shape=ft.RoundedRectangleBorder(radius=8),
                padding=ft.padding.symmetric(horizontal=18, vertical=12),
                text_style=ft.TextStyle(size=16),
            ),
            on_click=self.clear_filters
        )

        # Tabla de miembros
        self.members_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Nombre", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Email", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Teléfono", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Membresía", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Estado", size=18, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Acciones", size=18, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=12,
            vertical_lines=ft.border.all(1, ft.colors.GREY_300),
            horizontal_lines=ft.border.all(1, ft.colors.GREY_300),
            column_spacing=60,
            heading_row_color=ft.colors.GREY_100,
            heading_row_height=60,
            data_row_color=ft.colors.WHITE,
            data_row_min_height=56,
        )

        # Modal de nuevo/editar miembro
        self.is_editing = False
        self.editing_member_id = None
        self.new_member_modal = ft.AlertDialog(
            title=ft.Text("Registrar Nuevo Miembro", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Text("Información Personal", size=16, weight=ft.FontWeight.BOLD),
                        ft.Row([
                            ft.TextField(
                                label="Nombre",
                                width=240,
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_name
                            ),
                            ft.TextField(
                                label="Apellido",
                                width=240,
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_lastname
                            ),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(height=8),
                        ft.Row([
                            ft.TextField(
                                label="Documento",
                                width=240,
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_document
                            ),
                            ft.TextField(
                                label="Email",
                                width=240,
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_email
                            ),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(height=8),
                        ft.Row([
                            ft.ElevatedButton(
                                self.get_birth_date_btn_text(),
                                icon=ft.icons.CALENDAR_TODAY,
                                on_click=self.show_birth_date_picker,
                                ref=self.new_member_birth_date,
                                style=ft.ButtonStyle(padding=ft.padding.symmetric(horizontal=12, vertical=10)),
                            ),
                            ft.Dropdown(
                                label="Género",
                                width=240,
                                options=[
                                    ft.dropdown.Option("Masculino"),
                                    ft.dropdown.Option("Femenino"),
                                    ft.dropdown.Option("Otro")
                                ],
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_gender
                            ),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(height=16),
                        ft.Text("Información de Contacto", size=16, weight=ft.FontWeight.BOLD),
                        ft.Row([
                            ft.TextField(
                                label="Teléfono",
                                width=240,
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_phone
                            ),
                            ft.TextField(
                                label="Dirección",
                                width=240,
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_address
                            ),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(height=16),
                        ft.Text("Información de Membresía", size=16, weight=ft.FontWeight.BOLD),
                        ft.Row([
                            ft.Dropdown(
                                label="Tipo de Membresía",
                                width=240,
                                options=[
                                    ft.dropdown.Option("Mensual"),
                                    ft.dropdown.Option("Trimestral"),
                                    ft.dropdown.Option("Semestral"),
                                    ft.dropdown.Option("Anual")
                                ],
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_membership
                            ),
                            ft.ElevatedButton(
                                self.get_start_date_btn_text(),
                                icon=ft.icons.CALENDAR_TODAY,
                                on_click=self.show_start_date_picker,
                                ref=self.new_member_start_date,
                                style=ft.ButtonStyle(padding=ft.padding.symmetric(horizontal=12, vertical=10)),
                            ),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(height=16),
                        ft.Text("Estado del Miembro", size=16, weight=ft.FontWeight.BOLD),
                        ft.Row([
                            ft.Dropdown(
                                label="Estado",
                                width=240,
                                options=[
                                    ft.dropdown.Option("Activo"),
                                    ft.dropdown.Option("Inactivo")
                                ],
                                border_radius=8,
                                text_size=16,
                                ref=self.new_member_status,
                                disabled=not self.is_editing,  # Solo habilitado en modo edición
                                value="Activo"  # Valor por defecto
                            ),
                        ], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(height=16),
                        ft.Text("Información Médica", size=16, weight=ft.FontWeight.BOLD),
                        ft.TextField(
                            label="Información Médica",
                            width=480,
                            border_radius=8,
                            text_size=16,
                            ref=self.new_member_medical,
                            multiline=True,
                            min_lines=3,
                            max_lines=5,
                            helper_text="Ingrese cualquier información médica relevante del miembro"
                        ),
                    ],
                    spacing=0,
                ),
                width=540,
                padding=20,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_modal),
                ft.ElevatedButton(
                    lambda: "Actualizar" if self.is_editing else "Guardar",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE,
                        color=ft.colors.WHITE,
                    ),
                    on_click=self.save_member,
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        # Modal de exportación
        self.export_dialog = ft.AlertDialog(
            title=ft.Text("Confirmar Exportación", size=22, weight=ft.FontWeight.BOLD),
            content=ft.Text(
                "¿Deseas exportar los miembros?\n"
                "El archivo se guardará en tu carpeta de Descargas.",
                size=16
            ),
            actions=[
                ft.TextButton(
                    "Cancelar",
                    on_click=self.close_export_dialog,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.WHITE,
                        color=ft.colors.BLACK87,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
                ft.ElevatedButton(
                    "Exportar",
                    on_click=self.confirm_export,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE_900,
                        color=ft.colors.WHITE,
                        shape=ft.RoundedRectangleBorder(radius=8),
                        padding=ft.padding.symmetric(horizontal=28, vertical=12),
                        text_style=ft.TextStyle(size=16, weight=ft.FontWeight.BOLD)
                    )
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
            modal=True,
        )
        self.page.overlay.append(self.export_dialog)

    def get_content(self):
        content = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Row(
                        controls=[
                            self.welcome_title,
                            ft.Row(
                                controls=[
                                    self.new_member_btn,
                                    self.export_excel_btn,
                                    self.export_pdf_btn,
                                ],
                                alignment=ft.MainAxisAlignment.END,
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    ),
                    ft.Container(height=20),
                    ft.Row(
                        controls=[
                            self.search_field,
                            self.status_filter,
                            self.membership_type,
                            self.clear_btn,
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    ),
                    ft.Container(height=20),
                    ft.Column(
                        controls=[self.members_table],
                        scroll=ft.ScrollMode.AUTO,
                        expand=True,
                    ),
                ],
                spacing=0,
                expand=True,
            ),
            padding=20,
            expand=True,
        )
        self.page.update()
        return content

    def load_data(self):
        """
        Carga los datos de miembros
        """
        try:
            members = self.member_controller.get_members()
            self.update_members_table(members)
        except Exception as e:
            print(f"Error al cargar datos: {str(e)}")
            self.show_message(f"Error al cargar los datos: {str(e)}", ft.colors.RED)
            # Intentar reconectar
            try:
                self.db_session.rollback()
            except:
                pass

    def update_members_table(self, miembros):
        """
        Actualiza la tabla con los miembros
        """
        self.members_table.rows.clear()
        for member in miembros:
            # Crear la celda de acciones
            action_buttons = [
                ft.IconButton(
                    icon=ft.icons.EDIT,
                    icon_color=ft.colors.BLUE,
                    tooltip="Editar",
                    on_click=lambda e, m=member: self.edit_member(m)
                ),
                ft.IconButton(
                    icon=ft.icons.DELETE,
                    icon_color=ft.colors.RED,
                    tooltip="Eliminar",
                    on_click=lambda e, m=member: self.delete_member(m)
                ),
            ]
            
            # Verificar si el miembro tiene una rutina asignada
            has_routine = member.id_rutina is not None
            
            # Agregar botón de ver rutina si tiene una asignada
            if has_routine:
                action_buttons.append(
                    ft.IconButton(
                        icon=ft.icons.VISIBILITY,
                        icon_color=ft.colors.GREEN,
                        tooltip="Ver rutina",
                        on_click=lambda e, m=member: self.view_member_routine(m)
                    )
                )
            
            # Agregar botón de asignar rutina
            action_buttons.append(
                ft.IconButton(
                    icon=ft.icons.FITNESS_CENTER,
                    icon_color=ft.colors.PURPLE,
                    tooltip="Asignar Rutina",
                    on_click=lambda e, m=member: self.show_assign_routine_modal(m)
                )
            )
            
            self.members_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(f"{member.nombre} {member.apellido}")),
                        ft.DataCell(ft.Text(member.correo_electronico)),
                        ft.DataCell(
                            ft.Row([
                                ft.Text(member.telefono or "-"),
                                ft.IconButton(
                                    icon=ft.icons.PHONE_ANDROID,
                                    icon_color=ft.colors.GREEN,
                                    tooltip="Abrir WhatsApp",
                                    on_click=lambda e, m=member: self.open_whatsapp(m)
                                ) if member.telefono else None,
                            ])
                        ),
                        ft.DataCell(ft.Text(member.tipo_membresia)),
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(
                                    "Activo" if member.estado else "Inactivo",
                                    color=ft.colors.WHITE
                                ),
                                bgcolor=ft.colors.GREEN if member.estado else ft.colors.RED,
                                border_radius=8,
                                padding=ft.padding.symmetric(horizontal=12, vertical=6),
                            )
                        ),
                        ft.DataCell(
                            ft.Row(action_buttons)
                        ),
                    ]
                )
            )
        self.page.update()

    def show_new_member_modal(self, e):
        self.is_editing = False
        self.editing_member_id = None
        self.clear_member_fields()
        self.new_member_modal.title = ft.Text("Registrar Nuevo Miembro", size=26, weight=ft.FontWeight.BOLD)
        self.new_member_modal.actions[1].text = "Guardar"
        self.new_member_status.current.disabled = True  # Deshabilitar el dropdown en nuevo miembro
        self.new_member_status.current.value = "Activo"  # Establecer valor por defecto
        if self.new_member_modal not in self.page.overlay:
            self.page.overlay.append(self.new_member_modal)
        self.new_member_modal.open = True
        self.page.update()

    def close_modal(self, e):
        """
        Cierra el modal y limpia los campos
        """
        self.new_member_modal.open = False
        self.page.update()

    def show_birth_date_picker(self, e):
        """
        Muestra el selector de fecha de nacimiento
        """
        self.birth_date_picker.open = True
        self.page.update()

    def show_start_date_picker(self, e):
        """
        Muestra el selector de fecha de inicio
        """
        self.start_date_picker.open = True
        self.page.update()

    def on_birth_date_change(self, e):
        """
        Maneja el cambio de fecha de nacimiento
        """
        if self.birth_date_picker.value:
            self.new_member_birth_date.current.text = self.birth_date_picker.value.strftime("%d/%m/%Y")
            self.page.update()

    def on_start_date_change(self, e):
        """
        Maneja el cambio de fecha de inicio de membresía
        """
        if self.start_date_picker.value:
            self.new_member_start_date.current.text = self.start_date_picker.value.strftime("%d/%m/%Y")
            self.page.update()

    def get_birth_date_btn_text(self):
        if self.birth_date_picker.value:
            return self.birth_date_picker.value.strftime("%d/%m/%Y")
        return "Fecha de Nacimiento"

    def get_start_date_btn_text(self):
        if self.start_date_picker.value:
            return self.start_date_picker.value.strftime("%d/%m/%Y")
        return "Fecha de Inicio"

    def save_member(self, e):
        """
        Guarda un nuevo miembro o actualiza uno existente
        """
        try:
            # Validar campos requeridos
            required_fields = {
                "nombre": self.new_member_name.current.value,
                "apellido": self.new_member_lastname.current.value,
                "documento": self.new_member_document.current.value,
                "email": self.new_member_email.current.value,
                "fecha_nacimiento": self.birth_date_picker.value,
                "genero": self.new_member_gender.current.value,
                "tipo_membresia": self.new_member_membership.current.value,
            }

            # Recolectar campos faltantes
            missing_fields = []
            for field, value in required_fields.items():
                if not value:
                    # Convertir el nombre del campo a un formato más amigable
                    field_name = field.replace('_', ' ').title()
                    missing_fields.append(field_name)

            # Si hay campos faltantes, mostrar mensaje
            if missing_fields:
                error_message = "Por favor complete los siguientes campos:\n" + "\n".join(f"• {field}" for field in missing_fields)
                self.show_message(error_message, ft.colors.RED)
                return

            # Validar formato de email
            if not "@" in required_fields["email"]:
                self.show_message("El email no tiene un formato válido", ft.colors.RED)
                return

            # Validar documento (solo números)
            if not required_fields["documento"].isdigit():
                self.show_message("El documento debe contener solo números", ft.colors.RED)
                return

            # Preparar datos del miembro
            member_data = {
                "nombre": required_fields["nombre"].strip(),
                "apellido": required_fields["apellido"].strip(),
                "documento": required_fields["documento"].strip(),
                "correo_electronico": required_fields["email"].strip(),
                "fecha_nacimiento": required_fields["fecha_nacimiento"],
                "genero": required_fields["genero"],
                "tipo_membresia": required_fields["tipo_membresia"],
                "telefono": self.new_member_phone.current.value.strip() if self.new_member_phone.current.value else None,
                "direccion": self.new_member_address.current.value.strip() if self.new_member_address.current.value else None,
                "informacion_medica": self.new_member_medical.current.value.strip() if self.new_member_medical.current.value else None,
            }

            # Agregar estado solo si estamos editando
            if self.is_editing:
                member_data["estado"] = self.new_member_status.current.value == "Activo"
            else:
                member_data["estado"] = True  # Por defecto, el miembro se crea activo

            if self.is_editing:
                # Actualizar miembro existente
                success, message = self.member_controller.update_member(self.editing_member_id, member_data)
                if success:
                    self.show_message("Miembro actualizado exitosamente", ft.colors.GREEN)
                    self.close_modal(e)
                    self.load_data()
                else:
                    self.show_message(f"Error al actualizar el miembro: {message}", ft.colors.RED)
            else:
                # Crear nuevo miembro
                success, message = self.member_controller.create_member(member_data)
                if success:
                    self.show_message("Miembro agregado exitosamente", ft.colors.GREEN)
                    self.close_modal(e)
                    self.load_data()
                else:
                    self.show_message(f"Error al crear el miembro: {message}", ft.colors.RED)

        except Exception as e:
            self.show_message(f"Error inesperado: {str(e)}", ft.colors.RED)

    def edit_member(self, member):
        self.is_editing = True
        self.editing_member_id = member.id_miembro
        self.new_member_modal.title = ft.Text("Editar Miembro", size=26, weight=ft.FontWeight.BOLD)
        self.new_member_modal.actions[1].text = "Actualizar"
        # Autocompletar campos
        self.new_member_name.current.value = member.nombre
        self.new_member_lastname.current.value = member.apellido
        self.new_member_document.current.value = member.documento
        self.new_member_email.current.value = member.correo_electronico
        self.birth_date_picker.value = member.fecha_nacimiento
        self.new_member_birth_date.current.text = self.get_birth_date_btn_text()
        self.new_member_gender.current.value = member.genero
        self.new_member_phone.current.value = member.telefono or ""
        self.new_member_address.current.value = member.direccion or ""
        self.new_member_membership.current.value = member.tipo_membresia
        self.start_date_picker.value = member.fecha_registro if hasattr(member, 'fecha_registro') else None
        self.new_member_start_date.current.text = self.get_start_date_btn_text()
        self.new_member_medical.current.value = member.informacion_medica or ""
        self.new_member_status.current.disabled = False  # Habilitar el dropdown en edición
        self.new_member_status.current.value = "Activo" if member.estado else "Inactivo"
        if self.new_member_modal not in self.page.overlay:
            self.page.overlay.append(self.new_member_modal)
        self.new_member_modal.open = True
        self.page.update()

    def clear_member_fields(self):
        self.new_member_name.current.value = ""
        self.new_member_lastname.current.value = ""
        self.new_member_document.current.value = ""
        self.new_member_email.current.value = ""
        self.birth_date_picker.value = None
        self.new_member_birth_date.current.text = self.get_birth_date_btn_text()
        self.new_member_gender.current.value = None
        self.new_member_phone.current.value = ""
        self.new_member_address.current.value = ""
        self.new_member_membership.current.value = None
        self.start_date_picker.value = None
        self.new_member_start_date.current.text = self.get_start_date_btn_text()
        self.new_member_medical.current.value = ""
        self.new_member_status.current.value = None
        self.page.update()

    def delete_member(self, member):
        """
        Elimina un miembro o muestra un mensaje si no se puede eliminar
        """
        # Crear diálogo de confirmación
        confirm_dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Confirmar eliminación"),
            content=ft.Text(f"¿Está seguro que desea eliminar a {member.nombre} {member.apellido}?"),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: self._close_dialog(confirm_dialog)),
                ft.ElevatedButton(
                    "Eliminar",
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.RED,
                        color=ft.colors.WHITE,
                    ),
                    on_click=lambda e: self._confirm_delete_member(member, confirm_dialog)
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        if confirm_dialog not in self.page.overlay:
            self.page.overlay.append(confirm_dialog)
        confirm_dialog.open = True
        self.page.update()

    def _confirm_delete_member(self, member, dialog):
        """
        Confirma la eliminación del miembro
        """
        try:
            success, message = self.member_controller.delete_member(member.id_miembro)
            if success:
                self.show_message(message, ft.colors.GREEN)
                self.load_data()
            else:
                # Si el mensaje indica que tiene pagos asociados, mostrar en rojo
                if "tiene pagos asociados" in message:
                    self.show_message(message, ft.colors.RED)
                else:
                    self.show_message(message, ft.colors.ORANGE)
        except Exception as e:
            self.show_message(f"Error al eliminar el miembro: {str(e)}", ft.colors.RED)
        finally:
            dialog.open = False
            self.page.update()

    def view_member_routine(self, member):
        """
        Muestra los detalles de la rutina asignada al miembro
        """
        try:
            # Obtener la rutina del miembro usando el ID de rutina
            rutina = self.routine_controller.get_routine_by_id(member.id_rutina)
            if not rutina:
                self.show_message("No se encontró la rutina asignada", ft.colors.RED)
                return

            controls = [
                ft.Text("Información de la Rutina", size=16, weight=ft.FontWeight.BOLD),
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text("Nombre:", size=14, weight=ft.FontWeight.BOLD),
                            ft.Text(rutina.nombre, size=14),
                        ]),
                        ft.Row([
                            ft.Text("Dificultad:", size=14, weight=ft.FontWeight.BOLD),
                            ft.Text(rutina.nivel_dificultad, size=14),
                        ]),
                        ft.Row([
                            ft.Text("Descripción:", size=14, weight=ft.FontWeight.BOLD),
                            ft.Text(rutina.descripcion or "-", size=14),
                        ]),
                    ], spacing=10),
                    padding=10,
                ),
                ft.Divider(),
            ]

            # Previsualización o descarga de archivo
            if hasattr(rutina, 'documento_rutina') and rutina.documento_rutina:
                controls.append(ft.Text("Documento de Rutina", size=16, weight=ft.FontWeight.BOLD))
                controls.append(
                    ft.ElevatedButton(
                        "Previsualizar/Descargar Archivo",
                        icon=ft.icons.PICTURE_AS_PDF,
                        on_click=lambda e, r=rutina: self._preview_routine_file(r)
                    )
                )

            details_modal = ft.AlertDialog(
                title=ft.Text(f"Rutina de {member.nombre} {member.apellido}", size=26, weight=ft.FontWeight.BOLD),
                content=ft.Container(
                    content=ft.Column(controls=controls, spacing=0),
                    width=600,
                    padding=20,
                ),
                actions=[
                    ft.TextButton("Cerrar", on_click=lambda e: self._close_dialog(details_modal)),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )

            if details_modal not in self.page.overlay:
                self.page.overlay.append(details_modal)
            details_modal.open = True
            self.page.update()

        except Exception as e:
            print(f"[DEBUG] Error al ver rutina del miembro: {str(e)}")
            self.show_message(f"Error al ver la rutina: {str(e)}", ft.colors.RED)

    def _preview_routine_file(self, rutina):
        import tempfile, os, webbrowser
        # Intentar determinar extensión (por ahora PDF por defecto)
        ext = '.pdf'
        # Si quieres mejorar esto, puedes guardar el nombre original del archivo y extraer la extensión
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_file:
            tmp_file.write(rutina.documento_rutina)
            tmp_file.flush()
            webbrowser.open(tmp_file.name)

    def apply_filters(self, e):
        """
        Aplica los filtros seleccionados
        """
        filters = {}
        
        if self.search_field.value:
            filters['search'] = self.search_field.value
        
        if self.status_filter.value and self.status_filter.value != "Todos":
            filters['status'] = self.status_filter.value == "Activo"
        
        if self.membership_type.value and self.membership_type.value != "Todos":
            filters['membership_type'] = self.membership_type.value
        
        members = self.member_controller.get_members(filters)
        self.update_members_table(members)

    def clear_filters(self, e):
        """
        Limpia todos los filtros
        """
        self.search_field.value = ""
        self.status_filter.value = "Todos"
        self.membership_type.value = "Todos"
        self.page.update()
        
        # Recargar datos sin filtros
        self.load_data()

    def show_success_dialog(self, file_path: str):
        """
        Muestra un diálogo de éxito con un botón para abrir la carpeta
        """
        def open_folder(e):
            folder_path = os.path.dirname(file_path)
            if os.name == 'nt':  # Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # macOS y Linux
                subprocess.run(['xdg-open', folder_path])
            success_dialog.open = False
            self.page.update()

        success_dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.icons.CHECK_CIRCLE, color=ft.colors.GREEN, size=30),
                ft.Text("¡Descargado con éxito!", size=20, weight=ft.FontWeight.BOLD),
            ]),
            content=ft.Column([
                ft.Text(f"El archivo se ha guardado en:", size=14),
                ft.Text(file_path, size=14, color=ft.colors.GREY_700),
                ft.Container(height=20),
                ft.ElevatedButton(
                    "Abrir carpeta",
                    icon=ft.icons.FOLDER_OPEN,
                    on_click=open_folder,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE,
                        color=ft.colors.WHITE,
                    ),
                ),
            ], tight=True, spacing=10),
            actions=[
                ft.TextButton("Cerrar", on_click=lambda e: self.close_success_dialog(success_dialog)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        if success_dialog not in self.page.overlay:
            self.page.overlay.append(success_dialog)
        success_dialog.open = True
        self.page.update()

    def close_success_dialog(self, dialog):
        """
        Cierra el diálogo de éxito
        """
        dialog.open = False
        self.page.update()

    def export_to_excel(self, e):
        """
        Exporta los miembros actuales a un archivo Excel
        """
        try:
            self.export_type = "excel"  # Establecer tipo de exportación
            
            # Obtener los miembros filtrados actuales
            filters = {}
            
            if self.search_field.value:
                filters['name'] = self.search_field.value
            
            if self.status_filter.value and self.status_filter.value != "Todos":
                filters['status'] = self.status_filter.value

            members = self.member_controller.get_members(filters)
            
            if not members:
                self.show_message("No hay miembros para exportar", ft.colors.ORANGE)
                return

            # Actualizar el contenido del diálogo con el número de miembros
            self.export_dialog.content.value = f"¿Deseas exportar {len(members)} miembros a Excel?\nEl archivo se guardará en tu carpeta de Descargas."
            
            # Guardar los miembros en una variable de instancia para usarla en confirm_export
            self.members_to_export = members
            
            # Mostrar el diálogo
            self.export_dialog.open = True
            self.page.update()

        except Exception as e:
            self.show_message(f"Error al preparar la exportación: {str(e)}", ft.colors.RED)

    def export_to_pdf(self, e):
        """
        Exporta los miembros actuales a un archivo PDF
        """
        try:
            self.export_type = "pdf"  # Establecer tipo de exportación
            
            # Obtener los miembros filtrados actuales
            filters = {}
            
            if self.search_field.value:
                filters['name'] = self.search_field.value
            
            if self.status_filter.value and self.status_filter.value != "Todos":
                filters['status'] = self.status_filter.value

            members = self.member_controller.get_members(filters)
            
            if not members:
                self.show_message("No hay miembros para exportar", ft.colors.ORANGE)
                return

            # Actualizar el contenido del diálogo con el número de miembros
            self.export_dialog.content.value = f"¿Deseas exportar {len(members)} miembros a PDF?\nEl archivo se guardará en tu carpeta de Descargas."
            
            # Guardar los miembros en una variable de instancia para usarla en confirm_export
            self.members_to_export = members
            
            # Mostrar el diálogo
            self.export_dialog.open = True
            self.page.update()

        except Exception as e:
            self.show_message(f"Error al preparar la exportación: {str(e)}", ft.colors.RED)

    def confirm_export(self, e):
        """
        Confirma la exportación y genera el archivo
        """
        try:
            # Obtener la ruta de la carpeta de descargas
            downloads_path = str(Path.home() / "Downloads")

            # Usar los miembros guardados en export_to_excel/export_to_pdf
            members = getattr(self, 'members_to_export', [])
            if not members:
                self.show_message("No hay miembros para exportar", ft.colors.ORANGE)
                return

            # Determinar el tipo de exportación
            if self.export_type == "excel":
                self._export_to_excel(members, downloads_path)
            elif self.export_type == "pdf":
                self._export_to_pdf(members, downloads_path)
            else:
                self.show_message("Error: Tipo de exportación no válido", ft.colors.RED)
                return

            # Limpiar los miembros guardados
            self.members_to_export = None

            # Cerrar el diálogo de confirmación
            self.close_export_dialog(e)

        except Exception as e:
            self.show_message(f"Error al exportar: {str(e)}", ft.colors.RED)

    def close_export_dialog(self, e):
        """
        Cierra el diálogo de exportación
        """
        self.export_dialog.open = False
        self.export_type = None  # Resetear el tipo de exportación
        self.page.update()

    def _export_to_excel(self, members, downloads_path):
        """
        Exporta los miembros a Excel
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Miembros"

        # Estilos
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        date_format = 'dd/mm/yyyy'

        # Configurar ancho de columnas
        ws.column_dimensions['A'].width = 25  # Nombre
        ws.column_dimensions['B'].width = 25  # Apellido
        ws.column_dimensions['C'].width = 20  # Documento
        ws.column_dimensions['D'].width = 20  # Teléfono
        ws.column_dimensions['E'].width = 30  # Email
        ws.column_dimensions['F'].width = 20  # Fecha Registro
        ws.column_dimensions['G'].width = 20  # Estado
        ws.column_dimensions['H'].width = 25  # Rutina Asignada

        # Escribir encabezados
        headers = ["Nombre", "Apellido", "Documento", "Teléfono", "Email", "Fecha Registro", "Estado", "Rutina Asignada"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Escribir datos
        for row, member in enumerate(members, 2):
            # Nombre
            cell = ws.cell(row=row, column=1)
            cell.value = member.nombre
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Apellido
            cell = ws.cell(row=row, column=2)
            cell.value = member.apellido
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Documento
            cell = ws.cell(row=row, column=3)
            cell.value = member.documento
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Teléfono
            cell = ws.cell(row=row, column=4)
            cell.value = member.telefono
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Email
            cell = ws.cell(row=row, column=5)
            cell.value = member.correo_electronico if member.correo_electronico else ""
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Fecha Registro
            cell = ws.cell(row=row, column=6)
            if hasattr(member, 'fecha_registro') and member.fecha_registro:
                cell.value = member.fecha_registro
                cell.number_format = date_format
            else:
                cell.value = ""
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Estado
            cell = ws.cell(row=row, column=7)
            cell.value = "Activo" if member.estado else "Inactivo"
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color de fondo según estado
            if member.estado:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")

            # Rutina Asignada
            cell = ws.cell(row=row, column=8)
            if hasattr(member, 'rutina') and member.rutina:
                cell.value = member.rutina.nombre
            else:
                cell.value = "Sin rutina"
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Congelar la primera fila
        ws.freeze_panes = 'A2'

        # Generar nombre del archivo
        filename = f"miembros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(downloads_path, filename)

        # Guardar archivo
        wb.save(filepath)

        # Mostrar mensaje de éxito
        self.show_message(f"Archivo Excel guardado en: {filepath}", ft.colors.GREEN)

    def _export_to_pdf(self, members, downloads_path):
        """
        Exporta los miembros a PDF
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from datetime import datetime

            # Generar nombre del archivo
            filename = f"miembros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(downloads_path, filename)

            # Crear el documento PDF
            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(letter),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )

            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Centrado
            )

            # Contenido del PDF
            elements = []

            # Título
            title = Paragraph("Reporte de Miembros", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))

            # Información de filtros aplicados
            filter_info = []
            if self.search_field.value:
                filter_info.append(f"Nombre: {self.search_field.value}")
            if self.status_filter.value and self.status_filter.value != "Todos":
                filter_info.append(f"Estado: {self.status_filter.value}")

            if filter_info:
                filter_text = " | ".join(filter_info)
                filter_paragraph = Paragraph(f"Filtros aplicados: {filter_text}", styles["Normal"])
                elements.append(filter_paragraph)
                elements.append(Spacer(1, 20))

            # Datos de la tabla
            data = [["Nombre", "Apellido", "Documento", "Teléfono", "Email", "Fecha Registro", "Estado", "Rutina"]]
            
            for member in members:
                data.append([
                    member.nombre,
                    member.apellido,
                    member.documento,
                    member.telefono,
                    member.correo_electronico if member.correo_electronico else "",
                    member.fecha_registro.strftime("%d/%m/%Y") if hasattr(member, 'fecha_registro') and member.fecha_registro else "",
                    "Activo" if member.estado else "Inactivo",
                    member.rutina.nombre if hasattr(member, 'rutina') and member.rutina else "Sin rutina"
                ])

            # Crear la tabla
            table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.2*inch, 2*inch, 1.2*inch, 1*inch, 1.5*inch])
            
            # Estilo de la tabla
            table_style = TableStyle([
                # Encabezados
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Alineación específica
                ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # Nombre y Apellido
                ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Email
            ])

            # Agregar colores de fondo según estado
            for i, member in enumerate(members, 1):
                if member.estado:
                    table_style.add('BACKGROUND', (6, i), (6, i), colors.HexColor('#E2EFDA'))
                else:
                    table_style.add('BACKGROUND', (6, i), (6, i), colors.HexColor('#FFD9D9'))

            table.setStyle(table_style)
            elements.append(table)

            # Pie de página con fecha y hora
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=1
            )
            footer = Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                footer_style
            )
            elements.append(Spacer(1, 20))
            elements.append(footer)

            # Generar el PDF
            doc.build(elements)

            # Mostrar mensaje de éxito
            self.show_message(f"Archivo PDF guardado en: {filepath}", ft.colors.GREEN)

        except Exception as e:
            self.show_message(f"Error al generar PDF: {str(e)}", ft.colors.RED)

    def show_message(self, content, bgcolor: str):
        """
        Muestra un mensaje al usuario
        """
        self.page.snack_bar = ft.SnackBar(
            content=ft.Text(content, color=ft.colors.WHITE),
            bgcolor=bgcolor,
            duration=5000,  # 5 segundos
            action="Cerrar",
            action_color=ft.colors.WHITE,
        )
        self.page.snack_bar.open = True
        self.page.update()

    def open_whatsapp(self, member):
        """
        Abre WhatsApp con el número de teléfono del miembro
        """
        if member.telefono:
            # Eliminar cualquier carácter no numérico del teléfono
            phone = ''.join(filter(str.isdigit, member.telefono))
            # Asegurarse de que el número tenga el formato correcto para WhatsApp
            if phone.startswith('0'):
                phone = phone[1:]
            if not phone.startswith('54'):
                phone = '54' + phone
            # Abrir WhatsApp Web con el número
            import webbrowser
            webbrowser.open(f'https://wa.me/{phone}')

    def show_assign_routine_modal(self, member):
        """
        Muestra el modal para asignar una rutina a un miembro
        """
        self.selected_member = member
        # Obtener rutinas no asignadas
        rutinas_disponibles = self.routine_controller.get_routines({'unassigned': True})
        
        # Si el miembro ya tiene una rutina, obtenerla y agregarla a la lista
        if member.id_rutina:
            rutina_actual = self.routine_controller.get_routine_by_id(member.id_rutina)
            if rutina_actual:
                rutinas_disponibles.append(rutina_actual)
        
        self.assign_routine_dropdown = ft.Ref[ft.Dropdown]()
        self.assign_routine_modal = ft.AlertDialog(
            title=ft.Text(f"Asignar Rutina a {member.nombre} {member.apellido}", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                content=ft.Column([
                    ft.Dropdown(
                        label="Seleccionar Rutina",
                        ref=self.assign_routine_dropdown,
                        options=[ft.dropdown.Option(str(r.id_rutina), r.nombre) for r in rutinas_disponibles],
                        width=400,
                    ),
                ]),
                width=500,
                padding=20,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=self.close_assign_routine_modal),
                ft.ElevatedButton(
                    "Asignar",
                    icon=ft.icons.SAVE,
                    on_click=self.assign_routine,
                    style=ft.ButtonStyle(
                        bgcolor=ft.colors.BLUE,
                        color=ft.colors.WHITE,
                    ),
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        # Si el miembro ya tiene una rutina, seleccionarla por defecto
        if member.id_rutina:
            self.assign_routine_dropdown.current.value = str(member.id_rutina)
        
        if self.assign_routine_modal not in self.page.overlay:
            self.page.overlay.append(self.assign_routine_modal)
        self.assign_routine_modal.open = True
        self.page.update()

    def assign_routine(self, e):
        """
        Asigna una rutina al miembro seleccionado
        """
        rutina_id = self.assign_routine_dropdown.current.value
        if not rutina_id:
            self.show_message("Seleccione una rutina", ft.colors.RED)
            return
        
        # Usar el método específico para asignar rutina
        success, message = self.member_controller.assign_routine_to_member(
            self.selected_member.id_miembro,
            int(rutina_id)
        )
        
        if success:
            self.show_message("Rutina asignada exitosamente", ft.colors.GREEN)
            self.close_assign_routine_modal(e)
            self.load_data()  # Recargar los datos para actualizar la vista
        else:
            self.show_message(f"Error al asignar la rutina: {message}", ft.colors.RED)

    def view_member_routines(self, member):
        routines = self.routine_controller.get_member_routines(member.id_miembro)
        routines_list = ft.Column([
            ft.Text(f"Rutinas de {member.nombre} {member.apellido}", size=20, weight=ft.FontWeight.BOLD),
            ft.Container(height=20),
        ])
        if routines:
            for routine in routines:
                routines_list.controls.append(
                    ft.Card(
                        content=ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Icon(
                                        ft.icons.PICTURE_AS_PDF if routine.documento_rutina and routine.nombre and routine.nombre.lower().endswith('.pdf') else ft.icons.TABLE_VIEW,
                                        color=ft.colors.BLUE
                                    ),
                                    ft.Text(routine.nombre, size=16, weight=ft.FontWeight.BOLD),
                                ]),
                                ft.Text(f"Nivel: {routine.nivel_dificultad}"),
                                ft.Text(f"Descripción: {routine.descripcion or 'Sin descripción'}"),
                                ft.Text(f"Fecha: {routine.fecha_horario.strftime('%d/%m/%Y %H:%M') if routine.fecha_horario else '-'}"),
                                ft.Row([
                                    ft.ElevatedButton(
                                        "Descargar",
                                        icon=ft.icons.DOWNLOAD,
                                        on_click=lambda e, r=routine: self.download_routine(r)
                                    ),
                                    ft.IconButton(
                                        icon=ft.icons.DELETE,
                                        icon_color=ft.colors.RED,
                                        tooltip="Eliminar",
                                        on_click=lambda e, m=member, r=routine: self.delete_routine(m, r)
                                    ),
                                ]),
                            ]),
                            padding=20,
                        ),
                        margin=10,
                    )
                )
        else:
            routines_list.controls.append(
                ft.Text("No hay rutinas asignadas", size=16, color=ft.colors.GREY)
            )
        self.routines_modal = ft.AlertDialog(
            title=ft.Text("Rutinas Asignadas", size=26, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                content=routines_list,
                width=600,
                height=400,
                padding=20,
            ),
            actions=[
                ft.TextButton("Cerrar", on_click=lambda e: self.close_routines_modal()),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.routines_modal not in self.page.overlay:
            self.page.overlay.append(self.routines_modal)
        self.routines_modal.open = True
        self.page.update()

    def close_routines_modal(self):
        """
        Cierra el modal de rutinas
        """
        self.routines_modal.open = False
        self.page.update()

    def download_routine(self, routine):
        try:
            import tempfile, os, webbrowser
            with tempfile.TemporaryDirectory() as temp_dir:
                ext = '.pdf' if routine.nombre and routine.nombre.lower().endswith('.pdf') else '.xlsx'
                file_path = os.path.join(temp_dir, (routine.nombre or f"rutina_{routine.id_rutina}") + ext)
                with open(file_path, 'wb') as f:
                    f.write(routine.documento_rutina)
                webbrowser.open(file_path)
            self.show_message("Archivo descargado exitosamente", ft.colors.GREEN)
        except Exception as e:
            self.show_message(f"Error al descargar el archivo: {str(e)}", ft.colors.RED)

    def delete_routine(self, member, routine):
        """
        Elimina una rutina y recarga las rutinas del miembro
        """
        success, message = self.routine_controller.delete_routine(routine.id_rutina)
        if success:
            self.show_message(message, ft.colors.GREEN)
            # Después de eliminar, recargar la lista de rutinas para este miembro
            self.view_member_routines(member)
        else:
            self.show_message(message, ft.colors.RED)

    def close_assign_routine_modal(self, e=None):
        if hasattr(self, 'assign_routine_modal'):
            self.assign_routine_modal.open = False
            self.page.update()

    def _close_dialog(self, dialog):
        """
        Cierra un diálogo genérico y actualiza la página.
        """
        dialog.open = False
        self.page.update()

# Para probar directamente:
if __name__ == "__main__":
    ft.app(target=MembersView)
